"""Papel de trabajo en Excel. Formato unico para todas las skills.

Cuatro hojas siempre:
  Conclusion  - referencia, alcance, riesgos cubiertos, conclusion, firma
  Detalle     - los datos, con formulas vivas donde ayuden a la revision
  Traza       - origen de cada dato (fichero!hoja!celda)
  Excepciones - solo lo que no encaja, ordenado por severidad

Las formulas se dejan visibles (no valores calculados) en las lineas de total,
para que el revisor pueda ver que el papel cuadra sin volver a ejecutar nada.
"""

from __future__ import annotations

import os
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .excepciones import BLOQUEANTE, DOCUMENTAR, INFORMATIVA, RESOLVER, Excepcion
from .traza import RegistroTrazas

AZUL = "1F3864"
GRIS = "D9D9D9"
GRIS_CLARO = "F2F2F2"
ROJO = "FFC7CE"
AMBAR = "FFEB9C"
VERDE = "C6EFCE"

COLOR_SEVERIDAD = {BLOQUEANTE: ROJO, RESOLVER: AMBAR,
                   DOCUMENTAR: GRIS_CLARO, INFORMATIVA: GRIS_CLARO}

_BORDE = Border(*[Side(style="thin", color="BFBFBF")] * 4)
FORMATO_EUR = '#,##0.00'


def _cabecera(ws, fila: int, textos: list[str]) -> None:
    for j, t in enumerate(textos, start=1):
        c = ws.cell(row=fila, column=j, value=t)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = _BORDE


def _ancho(ws, anchos: dict[int, int]) -> None:
    for col, w in anchos.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _auto_ancho(ws, df: pd.DataFrame, fila_cab: int, maximo: int = 55) -> None:
    for j, col in enumerate(df.columns, start=1):
        largo = max([len(str(col))] + [len(str(v)) for v in df[col].head(200)] or [10])
        ws.column_dimensions[get_column_letter(j)].width = min(max(largo + 2, 10), maximo)


class PapelDeTrabajo:
    """Constructor del papel. Uso:

        p = PapelDeTrabajo("F-1", "Arrendamientos financieros", "ACME SL", 2025)
        p.alcance("...").riesgos(["R004"]).concluye("Sin excepciones", "LIMPIO")
        p.detalle(df, "Cuadro de cuotas", totales=["cuota", "interes"])
        p.trazas(registro); p.excepciones(lista)
        p.guardar("01-papeles/F-1 Arrendamientos.xlsx")
    """

    def __init__(self, ref: str, titulo: str, cliente: str, ejercicio: int,
                 preparado_por: str = "Asistido por dula-audit v1.0.0"):
        self.ref = ref
        self.titulo = titulo
        self.cliente = cliente
        self.ejercicio = ejercicio
        self.preparado_por = preparado_por
        self._alcance = ""
        self._riesgos: list[str] = []
        self._conclusion = ""
        self._estado = "EN CURSO"
        self._fundamento = ""
        self._horas = None
        self._detalles: list[tuple[str, pd.DataFrame, list[str]]] = []
        self._trazas: RegistroTrazas | None = None
        self._excepciones: list[Excepcion] = []
        self._indicadores: list[tuple[str, Any]] = []

    # -- construccion ------------------------------------------------------
    def alcance(self, texto: str) -> "PapelDeTrabajo":
        self._alcance = texto
        return self

    def riesgos(self, ids: list[str]) -> "PapelDeTrabajo":
        self._riesgos = list(ids)
        return self

    def fundamento(self, texto: str) -> "PapelDeTrabajo":
        self._fundamento = texto
        return self

    def indicador(self, nombre: str, valor: Any) -> "PapelDeTrabajo":
        self._indicadores.append((nombre, valor))
        return self

    def horas(self, h: float) -> "PapelDeTrabajo":
        self._horas = h
        return self

    def concluye(self, texto: str, estado: str = "CONCLUIDO") -> "PapelDeTrabajo":
        self._conclusion = texto
        self._estado = estado
        return self

    def detalle(self, df: pd.DataFrame, nombre: str = "Detalle",
                totales: list[str] | None = None) -> "PapelDeTrabajo":
        self._detalles.append((nombre, df, totales or []))
        return self

    def trazas(self, registro: RegistroTrazas) -> "PapelDeTrabajo":
        self._trazas = registro
        return self

    def excepciones(self, lista: list[Excepcion]) -> "PapelDeTrabajo":
        self._excepciones = list(lista)
        return self

    # -- escritura ---------------------------------------------------------
    def _hoja_conclusion(self, wb: Workbook) -> None:
        ws = wb.active
        ws.title = "Conclusion"
        _ancho(ws, {1: 26, 2: 95})
        ws["A1"] = f"{self.ref} - {self.titulo}"
        ws["A1"].font = Font(bold=True, size=14, color=AZUL)
        ws.merge_cells("A1:B1")

        filas = [
            ("Cliente", self.cliente),
            ("Ejercicio", self.ejercicio),
            ("Referencia del papel", self.ref),
            ("Preparado por", self.preparado_por),
            ("Revisado por", "[JUICIO-AUDITOR] - pendiente de firma del revisor"),
            ("Estado", self._estado),
            ("Riesgos cubiertos", ", ".join(self._riesgos) or "[JUICIO-AUDITOR] sin vincular"),
            ("Horas estimadas", self._horas if self._horas is not None else "n/d"),
        ]
        f = 3
        for k, v in filas:
            ws.cell(row=f, column=1, value=k).font = Font(bold=True)
            ws.cell(row=f, column=1).fill = PatternFill("solid", fgColor=GRIS_CLARO)
            ws.cell(row=f, column=2, value=v).alignment = Alignment(wrap_text=True,
                                                                   vertical="top")
            f += 1

        f += 1
        for titulo, texto in (("ALCANCE Y PROCEDIMIENTOS", self._alcance),
                              ("FUNDAMENTO DEL ENFOQUE", self._fundamento),
                              ("CONCLUSION", self._conclusion)):
            if not texto:
                continue
            c = ws.cell(row=f, column=1, value=titulo)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=AZUL)
            ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=2)
            f += 1
            cel = ws.cell(row=f, column=1, value=texto)
            cel.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=2)
            ws.row_dimensions[f].height = max(30, min(len(texto) // 100 * 14 + 30, 300))
            f += 2

        if self._indicadores:
            ws.cell(row=f, column=1, value="INDICADORES").font = Font(bold=True)
            f += 1
            for nombre, valor in self._indicadores:
                ws.cell(row=f, column=1, value=nombre)
                cel = ws.cell(row=f, column=2, value=valor)
                if isinstance(valor, (int, float)):
                    cel.number_format = FORMATO_EUR
                f += 1

        # semaforo del estado
        color = {"LIMPIO": VERDE, "CONCLUIDO": VERDE, "CON EXCEPCIONES": AMBAR,
                 "BLOQUEADO": ROJO, "EN CURSO": GRIS}.get(self._estado, GRIS)
        ws["B8"].fill = PatternFill("solid", fgColor=color)

    def _hoja_detalle(self, wb: Workbook) -> None:
        for nombre, df, totales in self._detalles:
            ws = wb.create_sheet(nombre[:31])
            if df.empty:
                ws["A1"] = "Sin registros."
                continue
            visible = df.drop(columns=[c for c in df.columns if str(c).startswith("_")],
                              errors="ignore")
            _cabecera(ws, 1, [str(c) for c in visible.columns])
            for i, (_, r) in enumerate(visible.iterrows(), start=2):
                for j, col in enumerate(visible.columns, start=1):
                    v = r[col]
                    if isinstance(v, pd.Timestamp):
                        v = v.to_pydatetime()
                    elif pd.isna(v):
                        v = None
                    c = ws.cell(row=i, column=j, value=v)
                    c.border = _BORDE
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        c.number_format = FORMATO_EUR
                if i % 2 == 0:
                    for j in range(1, len(visible.columns) + 1):
                        ws.cell(row=i, column=j).fill = PatternFill("solid",
                                                                   fgColor=GRIS_CLARO)
            # fila de totales con FORMULA VIVA (no valor): el revisor ve el cuadre
            if totales:
                fila_total = len(visible) + 2
                ws.cell(row=fila_total, column=1, value="TOTAL").font = Font(bold=True)
                for j, col in enumerate(visible.columns, start=1):
                    if str(col) in totales:
                        letra = get_column_letter(j)
                        c = ws.cell(row=fila_total, column=j,
                                    value=f"=SUM({letra}2:{letra}{len(visible) + 1})")
                        c.font = Font(bold=True)
                        c.number_format = FORMATO_EUR
                        c.fill = PatternFill("solid", fgColor=GRIS)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(len(visible.columns))}1"
            _auto_ancho(ws, visible, 1)

    def _hoja_traza(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Traza")
        cols = ["Concepto", "Valor", "Fichero", "Hoja/Pagina", "Celda/Fila/Clausula",
                "Confianza", "Nota", "Requiere revision"]
        _cabecera(ws, 1, cols)
        filas = self._trazas.filas() if self._trazas else []
        for i, r in enumerate(filas, start=2):
            for j, col in enumerate(cols, start=1):
                c = ws.cell(row=i, column=j, value=r.get(col))
                c.border = _BORDE
            if r.get("Requiere revision"):
                for j in range(1, len(cols) + 1):
                    ws.cell(row=i, column=j).fill = PatternFill("solid", fgColor=AMBAR)
        if not filas:
            ws["A2"] = "Sin trazas registradas."
        _ancho(ws, {1: 40, 2: 20, 3: 28, 4: 16, 5: 22, 6: 11, 7: 40, 8: 18})
        ws.freeze_panes = "A2"

    def _hoja_excepciones(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Excepciones")
        cols = ["Codigo", "Severidad", "Area", "Descripcion", "Importe", "Cuenta",
                "Origen", "Causa sugerida", "Accion", "Norma"]
        _cabecera(ws, 1, cols)
        from .excepciones import ORDEN_SEVERIDAD
        ordenadas = sorted(self._excepciones,
                           key=lambda e: (ORDEN_SEVERIDAD[e.severidad],
                                          -abs(e.importe or 0)))
        for i, exc in enumerate(ordenadas, start=2):
            fila = exc.fila()
            for j, col in enumerate(cols, start=1):
                c = ws.cell(row=i, column=j, value=fila.get(col))
                c.border = _BORDE
                c.alignment = Alignment(wrap_text=True, vertical="top")
                if col == "Importe" and fila.get(col) is not None:
                    c.number_format = FORMATO_EUR
            ws.cell(row=i, column=2).fill = PatternFill(
                "solid", fgColor=COLOR_SEVERIDAD[exc.severidad])
        if not ordenadas:
            ws["A2"] = "Sin excepciones. Todos los procedimientos concluyen sin salvedades."
        _ancho(ws, {1: 11, 2: 14, 3: 8, 4: 60, 5: 14, 6: 14, 7: 26, 8: 45, 9: 45, 10: 16})
        ws.freeze_panes = "A2"

    def guardar(self, path: str) -> str:
        os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
        wb = Workbook()
        self._hoja_conclusion(wb)
        self._hoja_detalle(wb)
        self._hoja_traza(wb)
        self._hoja_excepciones(wb)
        wb.save(path)
        return os.path.abspath(path)


def exporta_tablas(path: str, tablas: dict[str, pd.DataFrame]) -> str:
    """Export simple multi-hoja, para anexos y listados auxiliares."""
    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    for nombre, df in tablas.items():
        ws = wb.create_sheet(nombre[:31])
        visible = df.drop(columns=[c for c in df.columns if str(c).startswith("_")],
                          errors="ignore")
        if visible.empty:
            ws["A1"] = "Sin registros."
            continue
        _cabecera(ws, 1, [str(c) for c in visible.columns])
        for i, (_, r) in enumerate(visible.iterrows(), start=2):
            for j, col in enumerate(visible.columns, start=1):
                v = r[col]
                if isinstance(v, pd.Timestamp):
                    v = v.to_pydatetime()
                elif pd.isna(v):
                    v = None
                c = ws.cell(row=i, column=j, value=v)
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    c.number_format = FORMATO_EUR
        ws.freeze_panes = "A2"
        _auto_ancho(ws, visible, 1)
    wb.save(path)
    return os.path.abspath(path)
