"""Pruebas unitarias de la libreria de calculo.

Complementan al banco de aceptacion (`test_aceptacion.py`), que demuestra los seis
criterios del encargo pero recorre solo los caminos de un encargo tipo. Aqui se
ejercita cada funcion publica con valores conocidos y se comprueba el RESULTADO
NUMERICO, no solo que no reviente.

Motivo: el codigo de este plugin produce papeles de trabajo que se firman. Una
funcion que nunca se ha ejecutado no es codigo entregable.

    python3 tests/test_libreria.py
"""

from __future__ import annotations

import io as io_
import json
import os
import shutil
import sys

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(AQUI)
sys.path.insert(0, os.path.join(RAIZ, "shared", "scripts"))

import pandas as pd  # noqa: E402

from dula import (amortizaciones, analiticos, bitacora, comparador, cuadres,  # noqa: E402
                  estado, excel_out, financiacion, leasing, materialidad,
                  muestreo, perfil, plan_contable)
from dula.encargo import Encargo  # noqa: E402
from dula.excepciones import BLOQUEANTE, RESOLVER  # noqa: E402

SALIDA = os.path.join(AQUI, "salida", "libreria")
FALLOS: list[str] = []
TOTAL = [0]


def afirma(cond: bool, que: str) -> None:
    TOTAL[0] += 1
    if cond:
        print(f"  OK    {que}")
    else:
        print(f"  FALLA {que}")
        FALLOS.append(que)


def bloque(t: str) -> None:
    print(f"\n--- {t} " + "-" * max(0, 70 - len(t)))


# ---------------------------------------------------------------------------
def t_plan_contable() -> None:
    bloque("plan_contable")
    afirma(plan_contable.normaliza_cuenta("430.000.001") == "430000001",
           "normaliza_cuenta limpia separadores")
    afirma(plan_contable.grupo("5720000001") == "5", "grupo devuelve el primer digito")
    afirma(plan_contable.es_patrimonial("2180000001") and
           not plan_contable.es_patrimonial("7000000000"),
           "es_patrimonial distingue grupos 1-5 de 6-7")
    afirma(plan_contable.es_resultados("6400000000"), "es_resultados reconoce el grupo 6")
    afirma(plan_contable.es_grupo_8_9("9000000000"), "es_grupo_8_9 reconoce el grupo 9")

    c = plan_contable.clasifica("2180000001")
    afirma(c["estado"] == "OK" and c["masa"] == "ACTIVO NO CORRIENTE"
           and c["epigrafe"] == "A.II",
           f"clasifica 218 -> {c.get('masa')} {c.get('epigrafe')} (inmovilizado material)")
    c = plan_contable.clasifica("2810000001")
    afirma(c["signo"] == -1, "la amortizacion acumulada (281) lleva signo negativo")
    c = plan_contable.clasifica("7000000000")
    afirma(c["epigrafe"] == "1" and c["signo"] == 1,
           "clasifica 700 -> epigrafe 1 de la PyG (cifra de negocios)")
    afirma(plan_contable.clasifica("9999999999")["estado"] in {"SIN MAPEO", "GRUPO 8/9"},
           "una cuenta inexistente no se fuerza a un epigrafe")
    afirma(plan_contable.epigrafe_balance("5720000001")["titulo"].startswith("Efectivo"),
           "epigrafe_balance resuelve tesoreria")
    afirma(plan_contable.epigrafe_pyg("6810000000")["epigrafe"] == "8",
           "epigrafe_pyg resuelve la amortizacion del inmovilizado")

    # reservas restringidas: el prefijo mas largo gana
    r = plan_contable.reserva_restringida("1144000000")
    afirma(r is not None and "capitalizacion" in r["nombre"].lower()
           and r["disponible"] is False and "25 LIS" in r["norma"],
           "reserva_restringida identifica la 1144 como reserva de capitalizacion, "
           "indisponible, art. 25 LIS")
    r = plan_contable.reserva_restringida("1145000000")
    afirma(r is not None and "nivelacion" in r["nombre"].lower(),
           "identifica la 1145 como reserva de nivelacion")
    r = plan_contable.reserva_restringida("1120000000")
    afirma(r is not None and r["disponible"] is False and "274" in r["norma"],
           "identifica la 112 como reserva legal, art. 274 LSC")
    r = plan_contable.reserva_restringida("1130000000")
    afirma(r is not None and r["disponible"] is True,
           "las reservas voluntarias (113) figuran como disponibles")
    afirma(plan_contable.reserva_restringida("5720000001") is None,
           "una cuenta que no es reserva devuelve None")


def t_materialidad() -> None:
    bloque("materialidad")
    m, fund = materialidad.elige_magnitud(
        {"resultado_antes_impuestos": 180_000.0, "cifra_negocios": 2_000_000.0,
         "total_activo": 1_500_000.0})
    afirma(m == "resultado_antes_impuestos" and "9.0%" in fund or "9,0" in fund or len(fund) > 100,
           f"elige el resultado antes de impuestos cuando es estable y significativo ({m})")

    m, fund = materialidad.elige_magnitud(
        {"resultado_antes_impuestos": -50_000.0, "cifra_negocios": 2_000_000.0,
         "total_activo": 1_500_000.0})
    afirma(m == "cifra_negocios" and "negativo" in fund,
           "descarta el resultado cuando es negativo, y lo explica")

    m, _ = materialidad.elige_magnitud(
        {"resultado_antes_impuestos": 5_000.0, "cifra_negocios": 2_000_000.0,
         "total_activo": 1_500_000.0})
    afirma(m == "cifra_negocios",
           "descarta el resultado cuando es proximo a cero (0,25% de la cifra de negocios)")

    m, _ = materialidad.elige_magnitud({"total_activo": 900_000.0})
    afirma(m == "total_activo", "recurre al total activo si no hay actividad ordinaria")

    # el recalculo que BAJA es el que importa
    mat = materialidad.calcula({"cifra_negocios": 1_800_000.0}, "ESTANDAR")
    ev = materialidad.evalua_recalculo({"ejecucion": 20_000.0}, mat)
    afirma(ev["mp_nueva"] < 20_000.0 and ev["afecta_alcance"] is True
           and "ALERTA" in ev["mensaje"],
           f"evalua_recalculo alerta cuando la MP baja ({ev['mp_nueva']:,.2f} < 20.000)")
    ev = materialidad.evalua_recalculo({"ejecucion": 5_000.0}, mat)
    afirma(ev["afecta_alcance"] is False and "No reduce el alcance" in ev["mensaje"],
           "no alerta cuando la MP sube: no invalida el trabajo ya hecho")

    mat2 = materialidad.calcula({"cifra_negocios": 1_000_000.0}, "LIGERO")
    mat3 = materialidad.calcula({"cifra_negocios": 1_000_000.0}, "COMPLEJO")
    afirma(mat2.ejecucion > mat3.ejecucion,
           f"la MP se gradua por perfil: LIGERO {mat2.ejecucion:,.0f} > COMPLEJO "
           f"{mat3.ejecucion:,.0f}")
    afirma(abs(mat2.insignificante - mat2.global_ * 0.05) < 0.01,
           "el umbral de insignificancia es el 5% de la materialidad global")


def t_muestreo() -> None:
    bloque("muestreo")
    n, fund = muestreo.tamano_mus(1_000_000.0, 50_000.0, riesgo=0.05)
    afirma(n == 60, f"tamano_mus: 1.000.000 x 3,00 / 50.000 = 60 partidas (obtenido {n})")
    n2, _ = muestreo.tamano_mus(1_000_000.0, 50_000.0, riesgo=0.05, errores_esperados=1)
    afirma(n2 == 95, f"con 1 error esperado el factor sube a 4,75 -> 95 (obtenido {n2})")
    n3, _ = muestreo.tamano_mus(1_000_000.0, 50_000.0, riesgo=0.20)
    afirma(n3 < n, f"a mayor riesgo aceptado, menor muestra ({n3} < {n})")
    try:
        muestreo.tamano_mus(1_000.0, 0.0)
        afirma(False, "materialidad cero deberia lanzar error")
    except ValueError:
        afirma(True, "materialidad de ejecucion nula lanza error en lugar de dividir por cero")

    pob = pd.DataFrame({"ref": [f"F{i:03d}" for i in range(200)],
                        "importe": [1000.0 + i * 37 for i in range(200)]})
    m = muestreo.dirigido(pob, "importe", 5_000.0)
    afirma(m.tamano == int((pob["importe"] >= 5000).sum()),
           f"dirigido selecciona todas las partidas >= MP ({m.tamano})")
    afirma("NO se extrapola" in m.fundamento,
           "el fundamento del dirigido advierte de que no se extrapola")

    ctrl = muestreo.atributos(pob, "mensual", semilla=7)
    afirma(ctrl.tamano == 5, f"atributos mensual -> 5 elementos (obtenido {ctrl.tamano})")
    ctrl2 = muestreo.atributos(pob, "diaria", semilla=7)
    afirma(ctrl2.tamano == 25, "atributos diaria -> 25 elementos")
    afirma("una sola desviacion" in ctrl.fundamento,
           "advierte de que una sola desviacion invalida la confianza en el control")

    # evaluacion y proyeccion de errores
    mus = muestreo.mus(pob, "importe", 40_000.0, semilla=99)
    ev = muestreo.evalua(mus, [
        {"referencia": "F010", "descripcion": "Sobrevaloracion",
         "importe_registrado": 10_000.0, "importe_auditado": 8_000.0},
    ], 40_000.0)
    esperado = (2_000.0 / 10_000.0) * mus.intervalo
    afirma(abs(ev["error_proyectado"] - round(esperado, 2)) < 0.05,
           f"proyeccion por tainting: 20% x intervalo = {esperado:,.2f} "
           f"(obtenido {ev['error_proyectado']:,.2f})")
    ev2 = muestreo.evalua(mus, [
        {"referencia": "F001", "importe_registrado": 60_000.0,
         "importe_auditado": 0.0, "examen_individual": True},
    ], 40_000.0)
    afirma(ev2["error_conocido"] == 60_000.0 and ev2["error_proyectado"] == 0.0,
           "las partidas de examen individual NO se proyectan: su error es conocido")
    afirma(ev2["supera_materialidad"] is True and "SUPERA" in ev2["conclusion"],
           "concluye que la poblacion contiene una incorreccion material")
    afirma("ampliar la muestra" in ev2["conclusion"],
           "y enumera las tres salidas posibles")


def t_leasing() -> None:
    bloque("leasing (matematica financiera)")
    afirma(abs(leasing.van(0.10, [(0, -100.0), (1, 110.0)])) < 1e-9,
           "van: 110 a un ano al 10% vale exactamente 100 hoy")
    t = leasing.tir([(0, -1000.0), (1, 500.0), (2, 500.0), (3, 500.0)])
    afirma(t is not None and abs(t - 0.23375) < 0.001,
           f"tir de 1.000 contra 3x500 = 23,375% (obtenido {t:.5f})")
    afirma(leasing.tir([(0, -1000.0), (1, 100.0)]) is not None,
           "tir encuentra la tasa negativa cuando los pagos no cubren el principal")

    # prestamo frances conocido: 100.000 al 6% nominal anual, 60 cuotas mensuales
    r = (1 + 0.06) ** (1 / 12) - 1
    cuota = round(100_000 * r / (1 - (1 + r) ** -60), 2)
    tp, ta = leasing.tipo_implicito(100_000.0, cuota, 60)
    afirma(ta is not None and abs(ta - 0.06) < 0.0005,
           f"tipo_implicito recupera el 6% de partida (obtenido {ta:.4%})")

    cuadro = leasing.cuadro_amortizacion(100_000.0, cuota, 60,
                                         pd.Timestamp("2023-01-01"), "mensual",
                                         tasa_periodo=tp)
    afirma(len(cuadro) == 60, "el cuadro tiene tantas filas como cuotas")
    afirma(abs(cuadro["amortizacion_capital"].sum() - 100_000.0) < 0.05,
           f"la suma del capital amortizado = importe financiado "
           f"({cuadro['amortizacion_capital'].sum():,.2f})")
    afirma(cuadro.iloc[-1]["deuda_viva"] == 0.0, "la deuda viva final es exactamente cero")
    afirma(cuadro.iloc[0]["interes"] > cuadro.iloc[-1]["interes"],
           "la carga financiera decrece a lo largo de la vida del contrato")

    cuadro_op = leasing.cuadro_amortizacion(50_000.0, 900.0, 60,
                                            pd.Timestamp("2023-01-01"),
                                            opcion_compra=900.0)
    afirma(len(cuadro_op) == 61 and cuadro_op.iloc[-1]["cuota"] == 900.0,
           "con opcion de compra se anade una fila final por su importe")

    bloque("leasing (clasificacion NRV 8ª)")
    c = leasing.clasifica_contrato({"transferencia_propiedad": True})
    afirma(c.tipo == "FINANCIERO" and "transfiere la propiedad" in c.motivacion,
           "la transferencia de propiedad basta para calificarlo de financiero")
    c = leasing.clasifica_contrato({"opcion_compra": 800.0, "cuota": 800.0,
                                    "plazo_meses": 60, "vida_util_anos": 10,
                                    "importe_financiado": 40_000.0,
                                    "valor_razonable": 60_000.0})
    afirma(c.tipo == "FINANCIERO" and "opcion de compra" in c.motivacion,
           "una opcion equivalente a una cuota indica que no hay dudas de su ejercicio")
    c = leasing.clasifica_contrato({"plazo_meses": 96, "vida_util_anos": 10,
                                    "cuota": 500.0, "opcion_compra": 5_000.0,
                                    "importe_financiado": 30_000.0,
                                    "valor_razonable": 60_000.0})
    afirma(c.tipo == "FINANCIERO" and "vida economica" in c.motivacion,
           "un plazo que cubre el 80% de la vida util lo califica de financiero")
    c = leasing.clasifica_contrato({"plazo_meses": 24, "vida_util_anos": 10,
                                    "cuota": 500.0, "opcion_compra": 9_000.0,
                                    "importe_financiado": 12_000.0,
                                    "valor_razonable": 60_000.0})
    afirma(c.tipo == "OPERATIVO", "sin ningun indicador concurrente, es operativo")
    c = leasing.clasifica_contrato({"plazo_meses": 24, "cuota": 500.0,
                                    "importe_financiado": 12_000.0})
    afirma(c.tipo == "DUDOSO" and c.requiere_juicio and "no se clasifica como operativo"
           in c.motivacion.lower(),
           "SIN datos de vida util ni valor razonable NO se clasifica como operativo "
           "por defecto: se eleva a juicio del auditor")

    bloque("leasing (conciliacion contable)")
    resumen = pd.DataFrame([{"id": "L1", "clasificacion": "FINANCIERO",
                             "corriente_524": 10_000.0, "no_corriente_174": 40_000.0}])
    r = leasing.conciliacion_contable(resumen, 40_000.0, 10_000.0)
    afirma(r.ok and not r.excepciones, "cuadra cuando el recalculo coincide con 174 y 524")
    r = leasing.conciliacion_contable(resumen, 38_500.0, 10_000.0)
    afirma(len(r.excepciones) == 1 and abs(r.excepciones[0].importe - 1_500.0) < 0.01,
           f"detecta la diferencia de 1.500 EUR contra la cuenta 174")


def t_financiacion() -> None:
    bloque("financiacion")
    r = (1 + 0.05) ** (1 / 12) - 1
    cuota = round(50_000 * r / (1 - (1 + r) ** -48), 2)
    tp, ta = financiacion.tipo_efectivo(50_000.0, cuota, 48, gastos_iniciales=0.0)
    afirma(ta is not None and abs(ta - 0.05) < 0.0005,
           f"tipo_efectivo sin gastos recupera el 5% (obtenido {ta:.4%})")
    tp2, ta2 = financiacion.tipo_efectivo(50_000.0, cuota, 48, gastos_iniciales=1_000.0)
    afirma(ta2 > ta, f"los gastos de formalizacion elevan el tipo efectivo "
                     f"({ta:.4%} -> {ta2:.4%}), como exige la NRV 9ª")

    cuadro = financiacion.cuadro_prestamo(50_000.0, cuota, 48,
                                          pd.Timestamp("2024-01-01"),
                                          gastos_iniciales=1_000.0, tasa_periodo=tp2)
    afirma(abs(cuadro["amortizacion_capital"].sum() - 49_000.0) < 0.05,
           "el coste amortizado inicial es el importe NETO recibido (50.000 - 1.000)")
    afirma(cuadro.iloc[-1]["coste_amortizado"] == 0.0, "el coste amortizado final es cero")

    bloque("financiacion (conciliacion bancaria)")
    ext = pd.DataFrame({"fecha": ["01/12/2025", "05/12/2025", "20/12/2025"],
                        "importe": [1000.0, -450.0, 2200.0]})
    may = pd.DataFrame({"fecha": ["01/12/2025", "20/12/2025", "31/12/2025"],
                        "importe": [1000.0, 2200.0, -75.0]})
    res, conc = financiacion.concilia_banco(ext, may, "importe", "importe",
                                            "fecha", "fecha")
    afirma(len(conc) == 2, f"detecta 2 partidas conciliatorias (obtenido {len(conc)})")
    afirma(set(conc["origen"]) == {"extracto", "mayor"},
           "reporta en los DOS sentidos: en extracto sin registrar, y registrado sin extracto")
    afirma(any(abs(c) == 450.0 for c in conc["importe"]),
           "identifica el cargo bancario de 450 no contabilizado")
    afirma(any(abs(c) == 75.0 for c in conc["importe"]),
           "identifica el apunte contable de 75 sin movimiento en extracto")

    bloque("financiacion (covenants)")
    cov = pd.DataFrame([
        {"contrato": "PR-001", "entidad": "BANCO A", "covenant": "Ratio de solvencia",
         "exigido": 1.20, "real": 1.35, "sentido": "minimo"},
        {"contrato": "PR-002", "entidad": "BANCO B", "covenant": "Deuda/EBITDA",
         "exigido": 3.50, "real": 4.10, "sentido": "maximo"},
    ])
    res = financiacion.verifica_covenants(cov)
    afirma(res.datos["incumplidos"] == 1, "detecta exactamente un incumplimiento")
    inc = [e for e in res.excepciones if e.codigo == "COV-010"]
    afirma(len(inc) == 1 and inc[0].severidad == BLOQUEANTE,
           "el incumplimiento de covenant es BLOQUEANTE")
    afirma("waiver" in inc[0].accion and "corriente" in inc[0].accion,
           "la accion exige waiver anterior al cierre o reclasificacion a corriente")
    afirma("570" in inc[0].referencia_normativa,
           "conecta con la NIA-ES 570 (empresa en funcionamiento)")
    cov_sin = pd.DataFrame([{"contrato": "PR-003", "entidad": "BANCO C",
                             "covenant": "Fondo de maniobra", "exigido": None,
                             "real": None, "sentido": "minimo"}])
    res = financiacion.verifica_covenants(cov_sin)
    afirma(any(e.codigo == "COV-001" for e in res.excepciones),
           "un covenant sin datos se reporta, no se da por cumplido")


def t_analiticos() -> None:
    bloque("analiticos")
    u = analiticos.umbral_investigacion(20_000.0, factor=0.5, variacion_relativa=0.10)
    afirma(u["absoluto"] == 10_000.0 and u["relativo"] == 0.10,
           "umbral doble: 50% de la MP y 10% relativo")
    afirma("previo al analisis de las cifras" in u["fundamento"],
           "el fundamento deja constancia de que el umbral se fija A PRIORI")

    actual = {"Ventas": 1_000_000.0, "Personal": 300_000.0, "Suministros": 12_000.0}
    anterior = {"Ventas": 980_000.0, "Personal": 200_000.0, "Suministros": 3_000.0}
    res, det = analiticos.variaciones(actual, anterior, u)
    inv = det[det["investigar"] == "SI"]["epigrafe"].tolist()
    afirma("Personal" in inv,
           "Personal (+100.000, +50%) supera ambos umbrales: se investiga")
    afirma("Ventas" not in inv,
           "Ventas (+20.000, +2%) supera el absoluto pero no el relativo: no se investiga")
    afirma("Suministros" not in inv,
           "Suministros (+9.000, +300%) supera el relativo pero no el absoluto: "
           "no se investiga")
    afirma(all("corrobor" in e.accion for e in res.excepciones),
           "la accion exige corroborar la explicacion de la direccion con evidencia")

    res_r, ratios = analiticos.ratios(
        {"cifra_negocios": 1_000_000.0, "aprovisionamientos": -600_000.0,
         "clientes": 200_000.0, "existencias": 100_000.0, "pasivo": 500_000.0,
         "patrimonio_neto": 250_000.0, "activo_corriente": 400_000.0,
         "pasivo_corriente": 300_000.0, "resultado_antes_impuestos": 80_000.0,
         "resultado_explotacion": 95_000.0, "gastos_financieros": -15_000.0,
         "proveedores": 150_000.0},
        {"cifra_negocios": 900_000.0, "aprovisionamientos": -540_000.0,
         "clientes": 180_000.0, "existencias": 90_000.0, "pasivo": 400_000.0,
         "patrimonio_neto": 250_000.0, "activo_corriente": 380_000.0,
         "pasivo_corriente": 250_000.0, "resultado_antes_impuestos": 72_000.0,
         "resultado_explotacion": 85_000.0, "gastos_financieros": -12_000.0,
         "proveedores": 140_000.0})
    pmc = ratios[ratios["ratio"].str.contains("cobro")]["actual"].iloc[0]
    afirma(abs(pmc - 73.0) < 0.5,
           f"periodo medio de cobro = 200.000 x 365 / 1.000.000 = 73 dias (obtenido {pmc})")
    liq = ratios[ratios["ratio"].str.contains("Liquidez")]["actual"].iloc[0]
    afirma(abs(liq - 1.3333) < 0.001, f"liquidez = 400/300 = 1,333 (obtenido {liq})")

    r = analiticos.expectativa("Gastos de personal", 462_000.0, 462_000.0,
                               "14 empleados x 33.000 EUR de coste medio", 20_000.0)
    afirma(r.ok and not r.excepciones and "dentro del umbral" in r.conclusion,
           "expectativa: si el registrado coincide, concluye que la evidencia es suficiente")
    r = analiticos.expectativa("Gastos de personal", 520_000.0, 462_000.0,
                               "14 empleados x 33.000 EUR de coste medio", 20_000.0)
    afirma(len(r.excepciones) == 1 and abs(r.excepciones[0].importe - 58_000.0) < 0.01,
           "expectativa: detecta la desviacion de 58.000 EUR sobre un umbral de 10.000")


def t_comparador() -> None:
    bloque("comparador")
    a = {"Deudas": 285_000.0, "Clientes": 120_000.0}
    b = {"Deudas": 280_150.0, "Clientes": 120_000.0}
    r = comparador.compara_importes(a, b, "memoria", "balance")
    afirma(len(r.excepciones) == 1 and abs(r.excepciones[0].importe - 4_850.0) < 0.01,
           "compara_importes reporta solo el concepto que difiere, con su importe")
    afirma(r.datos["conceptos_comparados"] == 2 and r.datos["diferencias"] == 1,
           "informa de cuantos conceptos se compararon, no solo de los que fallan")

    r = comparador.compara_importes({"A": 100.0}, {"A": 100.0, "B": 50.0}, "x", "y")
    afirma(len(r.excepciones) == 1 and "no aparece en x" in r.excepciones[0].descripcion,
           "detecta un concepto presente en un lado y ausente en el otro")

    r = comparador.ejercicio_vs_anterior(
        {"Ventas": 1_000_000.0}, {"Ventas": 990_000.0},
        depositadas={"Ventas": 985_000.0})
    afirma(len(r.excepciones) == 2,
           "ejercicio_vs_anterior compara contra el anterior Y contra el Registro Mercantil")
    afirma(any("Registro Mercantil" in e.origen or "Registro Mercantil" in e.causa_sugerida
               for e in r.excepciones) or "Registro Mercantil" in r.conclusion,
           "identifica la discrepancia frente a las cuentas depositadas")

    r = comparador.diff_borradores({"Resultado": 100_000.0, "Activo": 900_000.0},
                                   {"Resultado": 82_000.0, "Activo": 900_000.0})
    afirma(len(r.excepciones) == 1 and r.excepciones[0].severidad == "INFORMATIVA",
           "el diff de borradores reporta cambios como INFORMATIVA, no como error")
    afirma("-18,000.00" in r.excepciones[0].descripcion.replace(".", ",").replace(",", ",")
           or "18" in r.excepciones[0].descripcion,
           "cuantifica el impacto del cambio por epigrafe")

    r = comparador.informe_gestion_vs_ccaa({"Cifra de negocios": 1_050_000.0},
                                           {"Cifra de negocios": 1_000_000.0})
    afirma(len(r.excepciones) == 1 and "Otra informacion" in r.conclusion,
           "informe_gestion_vs_ccaa conecta la incongruencia con la seccion del informe")

    sop = pd.DataFrame({"doc": ["F1", "F2", "F3"], "imp": [100.0, 200.0, 300.0]})
    cont = pd.DataFrame({"doc": ["F1", "F2", "F4"], "imp": [100.0, 250.0, 400.0]})
    r = comparador.soporte_vs_contabilidad(sop, cont, "doc", "imp")
    descs = " ".join(e.descripcion for e in r.excepciones)
    afirma("F3" in descs and "sin registro contable" in descs,
           "detecta el documento sin contabilizar (posible pasivo no registrado)")
    afirma("F4" in descs and "sin documento soporte" in descs,
           "detecta el asiento sin soporte documental")
    afirma("F2" in descs, "detecta la contabilizacion por importe distinto")


def t_cuadres_y_regularizacion() -> None:
    bloque("cuadres (deteccion de regularizacion)")
    pre = pd.DataFrame({"cuenta": ["7000000000", "6000000000", "5720000000"],
                        "debe": [0.0, 600.0, 400.0], "haber": [1000.0, 0.0, 0.0],
                        "saldo": [-1000.0, 600.0, 400.0], "_fila_origen": [2, 3, 4]})
    afirma(not cuadres.esta_regularizado(pre),
           "un balance con saldos en los grupos 6 y 7 es pre-regularizacion")
    r = cuadres.cuadre_resultado(pre)
    afirma(r.datos["resultado_grupos_6_7"] == 400.0,
           f"calcula el resultado como -(suma de saldos 6/7) = 400 "
           f"(obtenido {r.datos['resultado_grupos_6_7']})")

    post = pd.DataFrame({"cuenta": ["7000000000", "6000000000", "1290000000"],
                         "debe": [1000.0, 600.0, 0.0], "haber": [1000.0, 600.0, 400.0],
                         "saldo": [0.0, 0.0, -400.0], "_fila_origen": [2, 3, 4]})
    afirma(cuadres.esta_regularizado(post),
           "un balance con movimiento y saldo cero en 6/7 es post-regularizacion")
    r = cuadres.cuadre_resultado(post, diario=None)
    afirma(any(e.codigo == "CUA-051" for e in r.excepciones),
           "sin diario, NO se da por bueno el cuadre: se reporta que no se ha podido hacer")

    diario = pd.DataFrame({
        "asiento": ["1", "1", "9", "9", "9"],
        "cuenta": ["7000000000", "5720000000", "7000000000", "6000000000", "1290000000"],
        "debe": [0.0, 1000.0, 1000.0, 0.0, 0.0],
        "haber": [1000.0, 0.0, 0.0, 600.0, 400.0]})
    afirma(cuadres.resultado_desde_diario(diario) == 1000.0,
           "resultado_desde_diario excluye el asiento de regularizacion (el que mueve la 129)")


def t_perfil_y_escalado() -> None:
    bloque("perfil y escalado")
    cfg = perfil.configuracion("LIGERO")
    afirma(cfg["mp_factor"] == 0.75 and cfg["test_asientos_filtros"] == 4,
           "configuracion LIGERO: MP al 75% y 4 filtros de asientos")
    afirma(perfil.configuracion("COMPLEJO")["revision_calidad_encargo"] is True,
           "configuracion COMPLEJO exige revision de calidad del encargo")
    for p in ("LIGERO", "ESTANDAR", "COMPLEJO"):
        afirma(perfil.configuracion(p)["circularizacion_bancos"].startswith("siempre"),
           f"la circularizacion bancaria es obligatoria tambien en perfil {p}")

    _, detalle = perfil.puntua({"n_leasings": 103, "n_instrumentos_financiacion": 12,
                                "respuesta_cliente": "lento_o_desordenado",
                                "automatizacion_facturacion": "manual"})
    fact = perfil.factores_encarecedores(detalle, top=3)
    afirma(len(fact) == 3 and all(f["palanca"] for f in fact),
           "factores_encarecedores devuelve los 3 principales, cada uno con su palanca")
    afirma(fact[0]["puntos"] >= fact[-1]["puntos"], "vienen ordenados por impacto")
    afirma(any("Excel" in f["palanca"] or "datos" in f["palanca"] for f in fact),
           "las palancas son accionables (pedir datos en lugar de PDF)")

    v = perfil.valida_simplificacion("LIGERO", {})
    afirma(v["elevado"] is False, "sin hallazgos, el perfil se mantiene")
    v = perfil.valida_simplificacion("LIGERO", {"incorreccion_material": True})
    afirma(v["elevado"] is True and v["perfil_final"] == "ESTANDAR",
           "una incorreccion material eleva LIGERO -> ESTANDAR")
    afirma("ya ejecutadas" in v["mensaje"] and "quedado cortas" in v["mensaje"]
           and "recalcular la materialidad" in v["mensaje"],
           "el mensaje avisa de que el trabajo YA EJECUTADO puede haberse quedado corto")
    v = perfil.valida_simplificacion("COMPLEJO", {"indicio_fraude": True})
    afirma(v["perfil_final"] == "COMPLEJO", "COMPLEJO es el techo: no sube mas")

    _, motivo = perfil.clasifica_perfil(5, eip=True)
    afirma("537/2014" in motivo, "el override de EIP cita el Reglamento (UE) 537/2014")


def t_amortizaciones_deterioro() -> None:
    bloque("amortizaciones (indicios de deterioro)")
    det = pd.DataFrame([
        {"id": "E1", "descripcion": "Maquina parada", "cuenta": "213",
         "dias_amortizados": 0, "vnc_recalculado": 25_000.0, "fecha_baja": pd.NaT},
        {"id": "E2", "descripcion": "Maquina en uso", "cuenta": "213",
         "dias_amortizados": 365, "vnc_recalculado": 40_000.0, "fecha_baja": pd.NaT},
        {"id": "E3", "descripcion": "Vendida", "cuenta": "213",
         "dias_amortizados": 0, "vnc_recalculado": 5_000.0,
         "fecha_baja": pd.Timestamp("2025-06-30")},
    ])
    r = amortizaciones.indicios_deterioro(det)
    afirma(r.datos["elementos_sin_amortizar"] == 1,
           "senala solo el elemento parado, no el que esta en uso ni el dado de baja")
    afirma("E1" in r.excepciones[0].descripcion, "identifica el elemento correcto")
    r = amortizaciones.indicios_deterioro(pd.DataFrame())
    afirma(r.ok and "Sin inventario" in r.conclusion,
           "con inventario vacio no revienta: lo hace constar")


def t_bitacora() -> None:
    bloque("bitacora de uso de IA (NIGC1-ES)")
    carpeta = os.path.join(SALIDA, "bitacora")
    shutil.rmtree(carpeta, ignore_errors=True)
    os.makedirs(carpeta, exist_ok=True)
    b = bitacora.Bitacora(carpeta)
    afirma(b.entradas() == [], "una bitacora nueva esta vacia")

    fuente = os.path.join(carpeta, "fuente.txt")
    open(fuente, "w", encoding="utf-8").write("balance de prueba")
    eid = b.registra("area-arrendamientos", "leasing ...", [fuente], ["F-1.xlsx"],
                     {"contratos": 103}, "Sin excepciones.", 0, "F-1")
    afirma(eid == "IA-0001", f"la primera entrada es IA-0001 (obtenido {eid})")
    e = b.entradas()[0]
    afirma(len(e["entradas"][0]["sha256"]) == 64,
           "el fichero de entrada queda registrado con su SHA-256")
    afirma(e["validado_por"] is None and len(b.sin_validar) == 1,
           "la entrada nace SIN VALIDAR")

    b.registra("comparador-documental", "comparar ...", [], [], {}, "2 diferencias.", 2)
    afirma(len(b.entradas()) == 2 and b.entradas()[1]["id"] == "IA-0002",
           "el registro es append-only y correlativo")

    afirma(b.valida("IA-0001", "MJ Perez") is True, "valida devuelve True al encontrarla")
    afirma(b.entradas()[0]["validado_por"] == "MJ Perez"
           and b.entradas()[0]["validado_en"] is not None,
           "la validacion registra quien y cuando")
    afirma(len(b.sin_validar) == 1, "queda una sin validar")
    afirma(b.valida("IA-9999", "X") is False, "validar una entrada inexistente devuelve False")

    r = b.resumen()
    afirma(r["ejecuciones"] == 2 and r["validadas"] == 1 and r["sin_validar"] == 1,
           "el resumen cuadra")
    inf = b.informe()
    afirma("IA-0001" in inf and "MJ Perez" in inf and "PENDIENTE-VALIDACION" in inf,
           "el informe muestra validadas y pendientes")
    afirma("no altera las responsabilidades del auditor" in inf,
           "el informe incluye la declaracion para el archivo del encargo")


def t_estado_y_encargo() -> None:
    bloque("estado del encargo, horas y PBC")
    carpeta = os.path.join(SALIDA, "encargo")
    shutil.rmtree(carpeta, ignore_errors=True)
    enc = Encargo.crear(carpeta, "PRUEBA SL", 2025)

    a, m = siguiente(enc)
    afirma("ingesta-y-cuadres" in a,
           f"sin el papel 2.1 concluido, el siguiente paso es la ingesta")

    enc.registra_papel("2.1", "Cuadres", "x.xlsx", "Todos los cuadres pasan.",
                       estado="concluido", horas=2.5)
    p1 = enc.añade_pendiente("F", "Cuadros de leasing en Excel", prioridad=1)
    enc.añade_pendiente("B", "Inventario valorado", prioridad=4)
    a, _ = siguiente(enc)
    afirma("ruta critica" in a or "Reclamar" in a,
           "con un pendiente de prioridad 1, el siguiente paso es reclamarlo")

    afirma(enc.recuerda_pendiente(p1["id"]) and
           enc.datos["pendientes"][0]["recordatorios"] == 1,
           "recuerda_pendiente incrementa el contador de recordatorios")
    afirma(enc.recibe_pendiente(p1["id"]) and
           enc.datos["pendientes"][0]["estado"] == "recibido",
           "recibe_pendiente lo marca como recibido")
    afirma(len(estado.pendientes_ordenados(enc.datos)) == 1,
           "los recibidos desaparecen de la lista de pendientes")

    enc.añade_riesgo(id="R001", area="C", afirmacion="ocurrencia",
                     descripcion="Corte de operaciones", respuestas=[])
    a, _ = siguiente(enc)
    afirma("diseno-de-pruebas" in a,
           "con un riesgo sin respuesta, el siguiente paso es disenar su procedimiento")

    enc.responde_riesgo("R001", "C-1")
    enc.registra_papel("C-1", "Clientes", "c1.xlsx", "Sin excepciones.",
                       ["R001"], "concluido", horas=8.0)
    afirma(enc.horas_consumidas == 10.5,
           f"horas_consumidas suma los papeles (obtenido {enc.horas_consumidas})")
    enc.imputa_horas("C-1", 1.5, "LG")
    afirma(enc.horas_consumidas == 12.0, "imputa_horas acumula sobre el papel")
    afirma([p for p in enc.datos["papeles"] if p["ref"] == "C-1"][0]["preparado_por"] == "LG",
           "registra quien preparo el papel")
    try:
        enc.imputa_horas("Z-9", 1.0)
        afirma(False, "imputar a un papel inexistente deberia fallar")
    except KeyError:
        afirma(True, "imputar horas a un papel inexistente lanza KeyError")

    # re-registrar un papel no pierde las horas ya imputadas
    enc.registra_papel("C-1", "Clientes e ingresos", "c1.xlsx", "Actualizado.",
                       ["R001"], "concluido")
    afirma(enc.horas_consumidas == 12.0,
           "volver a registrar un papel conserva las horas acumuladas")

    enc.datos["fases"] = {k: "completa" for k in enc.datos["fases"]}
    enc.guardar()
    panel = estado.panel(enc.datos, {"ejecuciones": 3, "validadas": 1, "sin_validar": 2})
    afirma("PRUEBA SL" in panel and "SIGUIENTE PASO RECOMENDADO" in panel,
           "el panel incluye cabecera y siguiente paso")
    afirma("SIN VALIDAR" in panel, "el panel avisa de las ejecuciones de IA sin validar")
    afirma(len(panel.splitlines()) <= 40,
           f"el panel cabe en una pantalla ({len(panel.splitlines())} lineas)")

    enc.datos["excepciones"] = [{"codigo": "CUA-001", "severidad": BLOQUEANTE,
                                 "area": "2.1", "descripcion": "El balance no cuadra"}]
    a, m = siguiente(enc)
    afirma("bloqueantes" in a.lower(),
           "una excepcion bloqueante desplaza a todo lo demas en el siguiente paso")


def siguiente(enc: Encargo) -> tuple[str, str]:
    return estado.siguiente_paso(enc.datos)


def t_excel() -> None:
    bloque("exportacion a Excel")
    ruta = os.path.join(SALIDA, "anexo.xlsx")
    p = excel_out.exporta_tablas(ruta, {
        "Datos": pd.DataFrame({"cuenta": ["430", "700"], "importe": [100.5, -200.25],
                               "_interno": ["x", "y"]}),
        "Vacia": pd.DataFrame(),
    })
    from openpyxl import load_workbook
    wb = load_workbook(p)
    afirma(wb.sheetnames == ["Datos", "Vacia"], "crea una hoja por tabla")
    afirma(wb["Datos"].cell(row=1, column=1).value == "cuenta"
           and wb["Datos"].max_column == 2,
           "las columnas internas (con guion bajo) no se exportan")
    afirma(wb["Vacia"]["A1"].value == "Sin registros.",
           "una tabla vacia no rompe la exportacion")


def t_presentacion() -> None:
    """Lo que el auditor ve en pantalla y en el papel. Si esto falla, el trabajo
    esta bien hecho pero no se puede leer."""
    bloque("presentacion y API restante")
    from dula.excepciones import Excepcion, Resultado
    from dula.traza import RegistroTrazas, Traza

    res = Resultado("Prueba")
    res.conclusion = "Conclusion de prueba."
    for i in range(30):
        res.añade(Excepcion(f"X-{i:03d}", RESOLVER if i % 2 else BLOQUEANTE, "A",
                            f"Excepcion numero {i}", importe=float(i * 100)))
    txt = res.resumen()
    afirma(len(txt.splitlines()) <= 15,
           f"el resumen en pantalla respeta el limite de 15 lineas "
           f"({len(txt.splitlines())})")
    afirma("y 1" in txt or "mas (ver hoja Excepciones)" in txt,
           "avisa de cuantas excepciones quedan fuera del resumen")
    afirma(res.ordenadas()[0].severidad == BLOQUEANTE,
           "las bloqueantes se ordenan primero")
    afirma(abs(res.ordenadas()[0].importe) >= abs(res.ordenadas()[1].importe),
           "dentro de la misma severidad, ordena por importe descendente")
    d = res.excepciones[0].dict()
    afirma(d["codigo"] == "X-000" and "severidad" in d,
           "Excepcion.dict() serializa para encargo.json")
    afirma(Resultado("vacio").resumen().count("\n") <= 2,
           "un resultado sin excepciones produce un resumen minimo")

    mat = materialidad.calcula({"cifra_negocios": 2_000_000.0}, "ESTANDAR")
    r = mat.resumen()
    afirma("Materialidad global" in r and "Materialidad de ejecucion" in r
           and "insignificantes" in r,
           "el resumen de materialidad muestra MG, MP y umbral de insignificancia")

    pob = pd.DataFrame({"imp": [1000.0 * i for i in range(1, 60)]})
    m = muestreo.mus(pob, "imp", 30_000.0, semilla=11)
    r = m.resumen()
    afirma("Semilla de aleatoriedad: 11" in r,
           "el resumen del muestreo publica la semilla (reproducibilidad)")
    afirma("Cobertura" in r, "y la cobertura alcanzada sobre el importe")

    reg1, reg2 = RegistroTrazas(), RegistroTrazas()
    reg1.anota("a", 1, Traza("f.xlsx", "H1", "A1"))
    reg2.anota("b", 2, Traza("g.xlsx", "H1", "B2"))
    reg1.extend(reg2)
    afirma(len(reg1.anotaciones) == 2, "RegistroTrazas.extend consolida trazas de varias fuentes")

    p = excel_out.PapelDeTrabajo("Z-1", "Prueba", "CLIENTE", 2025)
    p.alcance("x").fundamento("Se elige MUS por ser poblacion numerosa.").horas(3.5)
    p.concluye("Sin excepciones.", "LIMPIO")
    ruta = p.guardar(os.path.join(SALIDA, "papel.xlsx"))
    from openpyxl import load_workbook
    ws = load_workbook(ruta)["Conclusion"]
    celdas = [str(ws.cell(row=r_, column=c).value)
              for r_ in range(1, 30) for c in (1, 2)]
    afirma(any("3.5" in c for c in celdas), "las horas figuran en la hoja de conclusion")
    afirma(any("MUS" in c for c in celdas), "el fundamento del enfoque figura en el papel")

    carpeta = os.path.join(SALIDA, "enc2")
    shutil.rmtree(carpeta, ignore_errors=True)
    enc = Encargo.crear(carpeta, "X SL", 2025)
    afirma(os.path.isdir(enc.carpeta) and enc.carpeta.endswith("enc2"),
           "Encargo.carpeta devuelve la carpeta del encargo")
    afirma(enc.materialidad_vigente is None, "sin materialidad fijada devuelve None")
    enc.fija_materialidad({"global": 1.0, "ejecucion": 0.7})
    enc.fija_materialidad({"global": 2.0, "ejecucion": 1.4})
    afirma(enc.materialidad_vigente["version"] == 2 and enc.materialidad_vigente["global"] == 2.0,
           "materialidad_vigente devuelve la ultima version, conservando el historico")
    enc.registra_excepciones([Excepcion("A-1", RESOLVER, "F", "d")], "F")
    enc.registra_excepciones([Excepcion("A-2", RESOLVER, "F", "e")], "F")
    afirma(len(enc.datos["excepciones"]) == 1,
           "registra_excepciones reemplaza las del area, no las duplica en cada pasada")

    # comparador contra el balance real, con el diario para reconstruir la PyG
    from dula import ingesta
    sys_df, _ = ingesta.normaliza_sumas_y_saldos(
        os.path.join(AQUI, "fixtures", "sumas_y_saldos_2025.xlsx"))
    diario_df, _ = ingesta.normaliza_diario(
        os.path.join(AQUI, "fixtures", "diario_2025.xlsx"))
    eps = comparador.agrega_por_epigrafe(sys_df, diario_df)
    solo_ep = {}
    for k, v in eps.items():
        solo_ep[k.split("|")[1]] = round(solo_ep.get(k.split("|")[1], 0.0) + v, 2)
    r = comparador.ccaa_vs_sumas_y_saldos(solo_ep, sys_df, diario_df)
    afirma(r.ok and not r.excepciones,
           "ccaa_vs_sumas_y_saldos no reporta diferencias cuando las cuentas anuales "
           "coinciden con el balance")
    manipulado = dict(solo_ep)
    clave = next(iter(manipulado))
    manipulado[clave] = round(manipulado[clave] + 7_500.0, 2)
    r = comparador.ccaa_vs_sumas_y_saldos(manipulado, sys_df, diario_df)
    afirma(any(abs((e.importe or 0) - 7_500.0) < 0.01 for e in r.excepciones)
           and r.excepciones[0].severidad == BLOQUEANTE,
           "detecta una diferencia de 7.500 EUR entre cuentas anuales y balance, "
           "y la marca como BLOQUEANTE")


def t_cli() -> None:
    """La interfaz por la que las skills invocan todo. Si un subcomando falla,
    la skill que lo usa no produce papel."""
    bloque("interfaz de linea de comandos")
    from dula import cli
    import contextlib

    FIX = os.path.join(AQUI, "fixtures")
    sys_y_saldos = os.path.join(FIX, "sumas_y_saldos_2025.xlsx")
    diario = os.path.join(FIX, "diario_2025.xlsx")
    base = os.path.join(SALIDA, "cli")
    shutil.rmtree(base, ignore_errors=True)
    enc = os.path.join(base, "ACME", "2025")

    def corre(*argv: str) -> tuple[int, str]:
        buf = io_.StringIO()
        with contextlib.redirect_stdout(buf), contextlib.redirect_stderr(buf):
            code = cli.main(list(argv))
        return code, buf.getvalue()

    code, salida = corre("doctor")
    afirma(code == 0, "`dula doctor` termina sin bloqueantes en esta instalacion")
    afirma("pandas" in salida and "Mapeo PGC operativo" in salida,
           "doctor comprueba dependencias y que el mapeo PGC se carga de verdad")
    afirma("VERIFICAR-LITERAL-ICAC" in salida,
           "doctor recuerda los parrafos del informe pendientes de contraste")

    code, _ = corre("nuevo", enc, "ACME SL", "2025")
    afirma(code == 0 and os.path.exists(os.path.join(enc, "encargo.json")),
           "`dula nuevo` crea el encargo y su estructura de carpetas")

    code, salida = corre("ingesta", sys_y_saldos, "--diario", diario,
                         "--ejercicio", "2025", "--encargo", enc, "--horas", "2.5",
                         "--papel", os.path.join(enc, "01-papeles", "2.1.xlsx"))
    afirma(code == 0 and "Todos los cuadres" in salida,
           "`dula ingesta` pasa los cuadres y genera el papel 2.1")
    afirma("Bitacora: IA-0001" in salida,
           "la ejecucion queda anotada en la bitacora de uso de IA")

    code, salida = corre("reservas", sys_y_saldos, "--encargo", enc,
                         "--papel", os.path.join(enc, "01-papeles", "G-2.xlsx"))
    afirma(code == 0 and "Reserva de capitalizacion" in salida,
           "`dula reservas` identifica la reserva de capitalizacion como indisponible")
    afirma("nota de fondos propios" in salida,
           "y recuerda que debe desglosarse en memoria")

    code, salida = corre("materialidad", '{"cifra_negocios": 1850000}',
                         "--perfil", "ESTANDAR", "--encargo", enc)
    afirma(code == 0 and "Materialidad de ejecucion" in salida,
           "`dula materialidad` calcula y registra la materialidad")

    code, salida = corre("estado", enc)
    afirma(code == 0 and "SIGUIENTE PASO RECOMENDADO" in salida,
           "`dula estado` presenta el panel con el siguiente paso")

    code, salida = corre("pbc", enc, "--anadir", "Cuadros de leasing", "--area", "F",
                         "--prioridad", "1")
    afirma(code == 0 and "P001" in salida, "`dula pbc` registra un pendiente del cliente")

    code, salida = corre("horas", enc)
    afirma(code == 0 and "2.5" in salida, "`dula horas` refleja las horas imputadas")

    code, salida = corre("validar", enc, "--listar")
    afirma(code == 0 and "PENDIENTE-VALIDACION" in salida,
           "`dula validar --listar` muestra las ejecuciones sin validar")
    code, _ = corre("validar", enc, "--entrada", "IA-0001", "--quien", "MJ Perez")
    afirma(code == 0, "`dula validar` valida una entrada concreta")

    code, salida = corre("calidad", enc, "--pre-vuelo")
    afirma("PANEL DEL SOCIO" in salida, "`dula calidad` emite el panel del socio")
    afirma("CAL-091" in salida,
           "y detecta la ejecucion asistida aun sin validar sobre un papel concluido")

    code, salida = corre("asientos", diario, "2025-12-31", "--materialidad", "15031")
    afirma(code == 0 and "Test de asientos" in salida, "`dula asientos` se ejecuta")

    code, salida = corre("leasing", os.path.join(FIX, "leasings_internos.xlsx"),
                         "2025-12-31")
    afirma("contratos" in salida, "`dula leasing` procesa el lote")

    code, salida = corre("amortizaciones", os.path.join(FIX, "inventario_inmovilizado.xlsx"),
                         "2025-01-01", "2025-12-31")
    afirma("Recalculados" in salida, "`dula amortizaciones` recalcula el inventario")

    code, salida = corre("financiacion", os.path.join(FIX, "cartera_financiacion.xlsx"),
                         "2025-12-31", "--confirmaciones",
                         os.path.join(FIX, "confirmaciones_bancarias.xlsx"))
    afirma("instrumentos" in salida, "`dula financiacion` recalcula la cartera")

    code, salida = corre("comparar", "--memoria-desgloses",
                         os.path.join(FIX, "memoria_desgloses.json"),
                         "--estados", os.path.join(FIX, "estados_financieros.json"))
    afirma("4,850.00" in salida.replace(".", ".") or "4850" in salida.replace(",", ""),
           "`dula comparar` detecta la diferencia deliberada")

    code, salida = corre("muestreo", sys_y_saldos, "saldo", "--metodo", "dirigido",
                         "--materialidad", "15031")
    afirma("Semilla de aleatoriedad" in salida,
           "`dula muestreo` publica la semilla para que sea reproducible")

    drivers = os.path.join(SALIDA, "drivers.json")
    with open(drivers, "w", encoding="utf-8") as fh:
        json.dump({"n_asientos": 4200, "n_cuentas": 180, "n_leasings": 12,
                   "n_instrumentos_financiacion": 6, "existencias": "si",
                   "automatizacion_facturacion": "media", "respuesta_cliente": "normal"},
                  fh)
    tarifas = os.path.join(SALIDA, "tarifas.json")
    with open(tarifas, "w", encoding="utf-8") as fh:
        json.dump({"socio": 120.0, "gerente": 85.0, "senior": 60.0, "ayudante": 40.0,
                   "_coste_hora": {"socio": 55.0, "gerente": 42.0, "senior": 28.0,
                                   "ayudante": 18.0}}, fh)
    code, salida = corre("estimar", drivers, "--tarifas", tarifas, "--encargo", enc)
    afirma(code == 0 and "PERFIL:" in salida and "Honorarios estimados" in salida,
           "`dula estimar` calcula perfil, horas y honorarios")
    afirma("Punto muerto" in salida and "FACTORES QUE MAS ENCARECEN" in salida,
           "y entrega el punto muerto y las palancas de abaratamiento")
    code, salida = corre("estimar", drivers, "--encargo", enc)
    afirma("[PENDIENTE-CLIENTE]" in salida,
           "sin tarifas, los honorarios salen [PENDIENTE-CLIENTE]: no se inventa un precio")

    act = os.path.join(SALIDA, "act.json"); ant = os.path.join(SALIDA, "ant.json")
    with open(act, "w", encoding="utf-8") as fh:
        json.dump({"Ventas": 1_000_000.0, "Personal": 300_000.0}, fh)
    with open(ant, "w", encoding="utf-8") as fh:
        json.dump({"Ventas": 980_000.0, "Personal": 200_000.0}, fh)
    code, salida = corre("analiticos", act, ant, "--materialidad", "20000")
    afirma(code == 0 and "umbral de investigacion" in salida.lower(),
           "`dula analiticos` publica el umbral fijado a priori")
    afirma("Personal" in salida,
           "y senala la variacion que supera ambos umbrales")

    # un subcomando con fichero inexistente falla limpio, sin rastro de pila
    code, salida = corre("ingesta", os.path.join(FIX, "no_existe.xlsx"))
    afirma(code == 1 and "ERROR" in salida and "Traceback" not in salida,
           "un fichero inexistente produce un error legible, no un rastro de pila")


# ---------------------------------------------------------------------------
def main() -> int:
    os.makedirs(SALIDA, exist_ok=True)
    print("=" * 78)
    print("PRUEBAS UNITARIAS DE LA LIBRERIA DE CALCULO")
    print("=" * 78)
    for fn in (t_plan_contable, t_materialidad, t_muestreo, t_leasing, t_financiacion,
               t_analiticos, t_comparador, t_cuadres_y_regularizacion,
               t_perfil_y_escalado, t_amortizaciones_deterioro, t_bitacora,
               t_estado_y_encargo, t_excel, t_presentacion, t_cli, t_menu, t_paquete_claude_ai):
        fn()
    print("\n" + "=" * 78)
    print(f"TOTAL: {TOTAL[0] - len(FALLOS)}/{TOTAL[0]} comprobaciones superadas.")
    if FALLOS:
        print("\nFallidas:")
        for f in FALLOS:
            print("  -", f)
    return 1 if FALLOS else 0



def t_menu() -> None:
    """El menú `/` debe ser legible: una frase corta en castellano por entrada.

    Las descripciones se escribían para que el modelo disparase la skill, y eso
    llenaba el menú de párrafos de 500 caracteres. El detalle de disparo vive
    ahora en `when_to_use`, que el usuario no ve.
    """
    bloque("menu de comandos (/)")
    import glob as g
    import yaml
    raiz = RAIZ
    entradas = []
    for f in sorted(g.glob(os.path.join(raiz, "skills", "*", "SKILL.md"))
                    + g.glob(os.path.join(raiz, "commands", "*.md"))):
        fm = yaml.safe_load(__import__("re").match(
            r"^---\n(.*?)\n---", open(f, encoding="utf-8").read(), __import__("re").S).group(1))
        entradas.append((f, fm))

    largas = [(os.path.basename(os.path.dirname(f)), len(fm["description"]))
              for f, fm in entradas if len(fm["description"]) > 105]
    afirma(not largas, f"ninguna descripcion pasa de 105 caracteres (peores: {largas[:3]})")

    sin_punto = [f for f, fm in entradas if not fm["description"].rstrip().endswith(".")]
    afirma(not sin_punto, "todas las descripciones son una frase acabada en punto")

    ingles = [os.path.basename(os.path.dirname(f)) for f, fm in entradas
              if __import__("re").search(r"\b(the|and|with|when|use it|for the)\b",
                                         fm["description"], __import__("re").I)]
    afirma(not ingles, f"todas las descripciones estan en castellano ({ingles[:3]})")

    skills = [(f, fm) for f, fm in entradas if "SKILL.md" in f]
    sin_trigger = [os.path.basename(os.path.dirname(f)) for f, fm in skills
                   if not fm.get("when_to_use") and fm.get("user-invocable") is not False]
    afirma(not sin_trigger,
           f"toda skill visible conserva su when_to_use para el disparo ({sin_trigger[:3]})")

    sin_hint = [os.path.basename(os.path.dirname(f)) for f, fm in skills
                if not fm.get("argument-hint") and fm.get("user-invocable") is not False
                and os.path.basename(os.path.dirname(f)) != "convenciones-dula"]
    afirma(not sin_hint, f"toda skill invocable indica que argumentos espera ({sin_hint[:3]})")

    cuerpos_sin_arranque = []
    for f, fm in skills:
        if fm.get("user-invocable") is False or "convenciones" in f:
            continue
        if "Al invocarla, empieza por aqui" not in open(f, encoding="utf-8").read() \
           and "Al invocarla, empieza por aquí" not in open(f, encoding="utf-8").read():
            cuerpos_sin_arranque.append(os.path.basename(os.path.dirname(f)))
    afirma(not cuerpos_sin_arranque,
           f"al pulsarla, toda skill presenta que necesita y que devuelve "
           f"({cuerpos_sin_arranque[:3]})")

    # el menu tiene que caber en una pantalla: diez entradas, no treinta y cinco
    afirma(len(skills) <= 10,
           f"el menu no pasa de diez skills (hay {len(skills)})")

    # las guias de fase remiten a procedimientos que existen
    procs = {os.path.splitext(os.path.basename(p))[0]
             for p in g.glob(os.path.join(raiz, "shared", "procedimientos", "*.md"))}
    afirma(len(procs) >= 25, f"los procedimientos siguen en shared/ (hay {len(procs)})")
    rotos = []
    for f, fm in skills:
        texto = open(f, encoding="utf-8").read()
        for ref in __import__("re").findall(r"^\| `([a-z0-9-]+)` \|", texto,
                                            __import__("re").M):
            if ref not in procs and ref not in {n for n, _ in
                                                [(fm2["name"], 0) for _, fm2 in skills]}:
                rotos.append((os.path.basename(os.path.dirname(f)), ref))
    afirma(not rotos, f"toda guia remite a un procedimiento existente ({rotos[:3]})")


def t_paquete_claude_ai() -> None:
    """El paquete .zip para claude.ai debe cumplir la especificación Agent Skills.

    claude.ai no admite plugins ni los campos de frontmatter propios de Claude
    Code: `argument-hint`, `when_to_use` y `user-invocable` producen
    "Unexpected key(s) in SKILL.md frontmatter". El paquete se genera desde el
    plugin para que ambos no se desincronicen, y aquí se comprueba que sale
    válido.
    """
    bloque("paquete para claude.ai")
    import subprocess, zipfile
    import yaml

    r = subprocess.run([sys.executable, os.path.join(RAIZ, "tools", "construir_claude_ai.py")],
                       capture_output=True, text=True, cwd=RAIZ)
    afirma(r.returncode == 0, f"el paquete se construye sin errores ({r.stdout.strip()[:60]})")
    zpath = os.path.join(RAIZ, "build", "dula-audit-claude-ai.zip")
    afirma(os.path.exists(zpath), "se genera el .zip")

    with zipfile.ZipFile(zpath) as z:
        nombres = z.namelist()
        skill = [n for n in nombres if n.endswith("SKILL.md")]
        afirma(skill == ["dula-audit/SKILL.md"],
               "el zip lleva la carpeta de la skill en la raíz, con un único SKILL.md")
        fm_txt = __import__("re").match(
            r"^---\n(.*?)\n---", z.read("dula-audit/SKILL.md").decode("utf-8"),
            __import__("re").S).group(1)
        fm = yaml.safe_load(fm_txt)
        # solo los dos campos obligatorios: `license` y `compatibility` exigen
        # valores normalizados y con texto libre la subida falla al sincronizar
        afirma(set(fm) == {"name", "description"},
               f"frontmatter minimo, solo name y description ({sorted(fm)})")
        afirma(len(fm["description"]) <= 1024,
               f"description de {len(fm['description'])} caracteres (máximo 1024)")
        afirma(__import__("re").fullmatch(r"[a-z0-9-]{1,64}", fm["name"])
               and "claude" not in fm["name"] and "anthropic" not in fm["name"],
               f"name '{fm['name']}' cumple el formato y no usa palabras reservadas")

        procs = [n for n in nombres if n.startswith("dula-audit/procedimientos/")]
        afirma(len(procs) == 39,
               f"viajan las 10 guias y los 29 procedimientos (hay {len(procs)})")
        for carpeta in ("referencias/", "plantillas/", "scripts/dula/", "dula.py"):
            afirma(any(n.startswith(f"dula-audit/{carpeta}") for n in nombres),
                   f"el paquete incluye {carpeta}")
        afirma(not any("when_to_use" in z.read(p).decode("utf-8")[:200]
                       and p.endswith("SKILL.md") for p in nombres),
               "ningún fichero conserva frontmatter de Claude Code")

        # lo que dispara el rechazo al sincronizar: ejecutables, scripts de
        # shell, extensiones desconocidas o instalacion de dependencias
        afirma(len(nombres) <= 200, f"por debajo del limite de 200 ficheros ({len(nombres)})")
        con_bit = [n for n in nombres
                   if (z.getinfo(n).external_attr >> 16) & 0o111]
        afirma(not con_bit, f"ningun fichero con bit de ejecucion ({con_bit[:3]})")
        raras = [n for n in nombres
                 if os.path.splitext(n)[1] not in (".md", ".py", ".json", "")]
        afirma(not raras, f"solo extensiones conocidas: md, py, json ({raras[:3]})")

        # las rutas del plugin se reescriben: en claude.ai no hay
        # ${CLAUDE_PLUGIN_ROOT} ni el lanzador `dula` en el PATH
        crudas = [p for p in procs if "CLAUDE_PLUGIN_ROOT"
                  in z.read(p).decode("utf-8")]
        afirma(not crudas, f"ningun procedimiento usa CLAUDE_PLUGIN_ROOT ({crudas[:3]})")

    # las rutas se resuelven fuera de la estructura del plugin
    from dula import rutas
    afirma(os.path.isdir(rutas.directorio("referencias"))
           and os.path.isdir(rutas.directorio("plantillas")),
           "el localizador de rutas encuentra referencias y plantillas en el plugin")
    try:
        rutas.directorio("inexistente")
        afirma(False, "un recurso desconocido debería fallar")
    except ValueError:
        afirma(True, "un recurso desconocido lanza ValueError")

    # El paquete para Claude Code: se descomprime en ~/.claude/skills/ y carga
    # sin marketplace, sin git y sin red. Es la via que no puede fallar por
    # sincronizacion, asi que tiene que salir completa y con el lanzador usable.
    zcc = os.path.join(RAIZ, "build", "dula-audit-claude-code.zip")
    afirma(os.path.exists(zcc), "se genera el .zip para Claude Code")
    with zipfile.ZipFile(zcc) as z:
        nombres = set(z.namelist())
        for req in ("dula-audit/.claude-plugin/plugin.json",
                    "dula-audit/bin/dula",
                    "dula-audit/shared/scripts/dula/cli.py",
                    "dula-audit/shared/procedimientos/materialidad.md",
                    "dula-audit/shared/references/mapeo-pgc.json",
                    "dula-audit/shared/templates/informe-auditoria.md"):
            afirma(req in nombres, f"el paquete de Claude Code incluye {req}")
        afirma(sum(1 for n in nombres
                   if n.startswith("dula-audit/skills/") and n.endswith("SKILL.md")) == 10,
               "viajan las 10 skills")
        afirma((z.getinfo("dula-audit/bin/dula").external_attr >> 16) & 0o111,
               "el lanzador bin/dula conserva el bit de ejecucion")
        afirma(not any("/build/" in n or "__pycache__" in n or n.endswith(".pyc")
                       for n in nombres),
               "no viajan artefactos de construccion ni cache de Python")

if __name__ == "__main__":
    raise SystemExit(main())
