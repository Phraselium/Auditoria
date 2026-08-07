"""Banco de pruebas: demuestra los seis criterios de aceptacion ejecutandolos.

    python3 tests/test_aceptacion.py

No usa pytest para no anadir dependencias al despacho: es un script autonomo que
devuelve codigo 0 si los seis criterios pasan.
"""

from __future__ import annotations

import json
import os
import shutil
import sys

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(AQUI)
sys.path.insert(0, os.path.join(RAIZ, "shared", "scripts"))

import pandas as pd  # noqa: E402

from dula import (amortizaciones, analiticos, asientos, calidad, comparador,  # noqa: E402
                  cuadres, financiacion, ingesta, leasing, materialidad,
                  muestreo, perfil)
from dula.encargo import Encargo  # noqa: E402
from dula.excel_out import PapelDeTrabajo  # noqa: E402
from dula.traza import RegistroTrazas, Traza, huella  # noqa: E402

FIXT = os.path.join(AQUI, "fixtures")
SALIDA = os.path.join(AQUI, "salida")
CIERRE = "2025-12-31"
EJERCICIO = 2025

RESULTADOS: list[tuple[str, bool, str]] = []


def comprueba(criterio: str, condicion: bool, detalle: str) -> None:
    RESULTADOS.append((criterio, condicion, detalle))
    marca = "OK  " if condicion else "FALLA"
    print(f"  [{marca}] {detalle}")


def titulo(t: str) -> None:
    print()
    print("=" * 78)
    print(t)
    print("=" * 78)


# ---------------------------------------------------------------------------
def criterio_1_encargo_completo() -> dict:
    """1. Un encargo pequeno se recorre de principio a fin sin salir del plugin."""
    titulo("CRITERIO 1 - Recorrido completo de un encargo pequeno")

    carpeta = os.path.join(SALIDA, "encargo_completo")
    if os.path.exists(carpeta):
        shutil.rmtree(carpeta)
    enc = Encargo.crear(carpeta, "TRANSPORTES SINTETICOS SL", EJERCICIO, "PGC-PYMES")
    enc.datos["cliente"]["nif"] = "B00000000"

    # --- Fase 0: perfil y estimacion ---
    drivers = {
        "n_asientos": 212, "n_cuentas": 34, "automatizacion_facturacion": "media",
        "n_instrumentos_financiacion": 5, "n_leasings": 103, "existencias": "no",
        "subvenciones": "no", "operaciones_vinculadas": "si",
        "moneda_extranjera": "no", "consolidacion": "no", "primer_encargo": "no",
        "respuesta_cliente": "normal",
    }
    puntuacion, detalle = perfil.puntua(drivers)
    p, motivo = perfil.clasifica_perfil(puntuacion)
    tarifas = {"socio": 120.0, "gerente": 85.0, "senior": 60.0, "ayudante": 40.0,
               "_coste_hora": {"socio": 55.0, "gerente": 42.0, "senior": 28.0,
                               "ayudante": 18.0}}
    est = perfil.estima(drivers, p, tarifas)
    enc.datos["perfil"] = {"puntuacion": puntuacion, "perfil": p, "motivo": motivo,
                           "drivers": drivers, "estimacion": est}
    enc.datos["fases"]["aceptacion"] = "completa"
    print(f"  Perfil: {p} ({puntuacion} pts) | horas {est['horas_totales']} h | "
          f"honorarios {est['honorarios']['total']:,.2f} EUR")
    comprueba("1", p in {"LIGERO", "ESTANDAR", "COMPLEJO"},
              f"Fase 0: perfil determinado ({p}) y honorarios estimados")

    # --- Fase 1: ingesta, materialidad, riesgos ---
    sys_df, meta = ingesta.normaliza_sumas_y_saldos(
        os.path.join(FIXT, "sumas_y_saldos_2025.xlsx"))
    diario_df, meta_d = ingesta.normaliza_diario(os.path.join(FIXT, "diario_2025.xlsx"))
    anterior_df, _ = ingesta.normaliza_sumas_y_saldos(
        os.path.join(FIXT, "sumas_y_saldos_2024.xlsx"))
    print(f"  Ingesta: {meta['registros']} cuentas (longitudes "
          f"{meta['longitudes_cuenta']}, grupos {''.join(meta['grupos'])}), "
          f"{meta_d['apuntes']} apuntes en {meta_d['asientos']} asientos")
    comprueba("1", max(meta["longitudes_cuenta"]) >= 10,
              f"Ingesta de cuentas de {max(meta['longitudes_cuenta'])} digitos")

    res_cuadres, ok = cuadres.ejecuta_todos(sys_df, diario_df, anterior_df, EJERCICIO)
    for r in res_cuadres:
        print(f"    - {r.concepto}: {r.conclusion[:80]}")
    comprueba("1", ok, "Todos los cuadres de integridad pasan (puerta de entrada)")

    # el balance esta tomado despues de la regularizacion: la cuenta de resultados
    # se reconstruye desde el diario, excluyendo el asiento de regularizacion
    epigrafes = comparador.agrega_por_epigrafe(sys_df, diario_df)
    inc = round(sum(v for k, v in epigrafes.items() if k.startswith("PYG|1|")), 2)
    total_activo = round(sum(v for k, v in epigrafes.items()
                             if k.startswith("ACTIVO")), 2)
    print(f"  Cifra de negocios {inc:,.2f} EUR | total activo {total_activo:,.2f} EUR")
    comprueba("1", inc > 0,
              "La cuenta de resultados se reconstruye pese a estar el balance "
              "post-regularizacion")
    mat = materialidad.calcula({"cifra_negocios": inc, "total_activo": total_activo},
                               perfil=p, magnitud="cifra_negocios")
    materialidad.especifica(mat, "M", "retribuciones al organo de administracion y "
                                      "operaciones con partes vinculadas, cuyo desglose es "
                                      "legalmente exigido con independencia de su importe",
                            0.25)
    enc.fija_materialidad(mat.dict())
    print(f"  Materialidad: MG {mat.global_:,.2f} EUR | MP {mat.ejecucion:,.2f} EUR")
    comprueba("1", mat.global_ > 0 and bool(mat.fundamento),
              "Materialidad determinada con fundamento redactado")

    enc.añade_riesgo(id="R001", area="C", afirmacion="ocurrencia",
                     descripcion="Reconocimiento de ingresos antes del devengo "
                                 "(presuncion de fraude, NIA-ES 240.25)",
                     espectro="alto", significativo=True, fraude=True)
    enc.añade_riesgo(id="R002", area="F", afirmacion="exactitud y valoracion",
                     descripcion="Clasificacion de arrendamientos y calculo de la "
                                 "carga financiera",
                     espectro="alto", significativo=True)
    enc.añade_riesgo(id="R003", area="E", afirmacion="integridad",
                     descripcion="Deuda financiera no registrada o garantias no "
                                 "desglosadas", espectro="medio")
    enc.añade_riesgo(id="R004", area="A", afirmacion="exactitud",
                     descripcion="Error en el calculo de la amortizacion del ejercicio",
                     espectro="bajo")
    enc.añade_riesgo(id="R005", area="2.8", afirmacion="ocurrencia",
                     descripcion="Elusion de los controles por la direccion mediante "
                                 "asientos manuales", espectro="alto",
                     significativo=True, fraude=True)
    enc.datos["fases"]["planificacion"] = "completa"

    for f in ("sumas_y_saldos_2025.xlsx", "diario_2025.xlsx"):
        ruta = os.path.join(FIXT, f)
        enc.registra_fuente(ruta, huella(ruta), f)

    reg = RegistroTrazas()
    reg.anota("Total debe", round(float(sys_df["debe"].sum()), 2),
              Traza(meta["fichero"], meta["hoja"], meta["columnas"].get("debe", "")))
    papel = PapelDeTrabajo("2.1", "Ingesta y cuadres de integridad",
                           "TRANSPORTES SINTETICOS SL", EJERCICIO)
    papel.alcance("Normalizacion y bateria de cuadres previa al trabajo de campo.")
    papel.riesgos([]).trazas(reg)
    papel.excepciones([e for r in res_cuadres for e in r.excepciones])
    papel.detalle(sys_df, "Sumas y saldos", ["debe", "haber", "saldo"])
    papel.concluye("Todos los cuadres pasan. Se puede continuar.", "LIMPIO")
    ruta_papel = papel.guardar(os.path.join(carpeta, "01-papeles", "2.1 Cuadres.xlsx"))
    enc.registra_papel("2.1", "Ingesta y cuadres", ruta_papel,
                       "Todos los cuadres de integridad pasan.", [], "concluido")

    # --- Fase 2: trabajo de campo ---
    # F - Arrendamientos
    dfs = [ingesta.lee_tabla(os.path.join(FIXT, f))[0] for f in
           ("leasings_banco_norte.xlsx", "leasings_caja_sur.xlsx",
            "leasings_internos.xlsx")]
    contratos = pd.concat([leasing.normaliza_lote(d) for d in dfs], ignore_index=True)
    contratos["_fila_origen"] = range(2, len(contratos) + 2)
    res_lea, resumen_lea, cuadros_lea = leasing.procesa_lote(contratos, CIERRE, "leasings")
    ruta_f1 = _papel_area("F-1", "Arrendamientos - recalculo y clasificacion", res_lea,
                          carpeta, {"Resumen": (resumen_lea, ["deuda_viva_cierre"]),
                                    "Por ejercicio": (
                                        leasing.cuadro_por_ejercicio(cuadros_lea),
                                        ["cuotas", "carga_financiera"])}, ["R002"])
    enc.registra_papel("F-1", "Arrendamientos", ruta_f1, res_lea.conclusion,
                       ["R002"], "concluido")

    # E - Financiacion
    df_fin, _ = ingesta.lee_tabla(os.path.join(FIXT, "cartera_financiacion.xlsx"))
    res_fin, resumen_fin = financiacion.procesa_cartera(df_fin, CIERRE, "cartera")
    df_conf, _ = ingesta.lee_tabla(os.path.join(FIXT, "confirmaciones_bancarias.xlsx"))
    res_conf = financiacion.seguimiento_confirmaciones(df_conf)
    res_fin.excepciones.extend(res_conf.excepciones)
    ruta_e1 = _papel_area("E-1", "Financiacion y tesoreria", res_fin, carpeta,
                          {"Cartera": (resumen_fin, ["deuda_viva", "corriente",
                                                     "no_corriente"])}, ["R003"])
    enc.registra_papel("E-1", "Financiacion y tesoreria", ruta_e1, res_fin.conclusion,
                       ["R003"], "concluido")

    # A - Inmovilizado
    df_inv, _ = ingesta.lee_tabla(os.path.join(FIXT, "inventario_inmovilizado.xlsx"))
    res_amo, det_amo = amortizaciones.recalcula(df_inv, "2025-01-01", "2025-12-31")
    ruta_a1 = _papel_area("A-1", "Inmovilizado - recalculo de amortizaciones", res_amo,
                          carpeta, {"Inventario": (det_amo, ["coste",
                                                             "dotacion_recalculada",
                                                             "dotacion_contable"])},
                          ["R004"])
    enc.registra_papel("A-1", "Inmovilizado", ruta_a1, res_amo.conclusion,
                       ["R004"], "concluido")

    # 2.8 - Test de asientos del diario
    res_dia, sel_dia = asientos.analiza(diario_df, CIERRE, mat.ejecucion, p)
    ruta_28 = _papel_area("2.8", "Test de asientos del diario", res_dia, carpeta,
                          {"Seleccionados": (sel_dia, ["importe"])}, ["R005"])
    enc.registra_papel("2.8", "Test de asientos del diario", ruta_28,
                       res_dia.conclusion, ["R005"], "concluido")

    # C - Ingresos: analitico mensual + muestreo
    res_men, det_men = analiticos.evolucion_mensual(diario_df, "70")
    clientes = diario_df[diario_df["cuenta"].str.startswith("43")].copy()
    m = muestreo.mus(clientes, "debe", mat.ejecucion, semilla=42)
    res_men.datos["muestra"] = {"tamano": m.tamano, "semilla": m.semilla,
                                "cobertura": round(m.cobertura, 4)}
    ruta_c1 = _papel_area("C-1", "Clientes e ingresos", res_men, carpeta,
                          {"Evolucion mensual": (det_men, ["importe"]),
                           "Muestra": (m.seleccion, ["_importe"])}, ["R001"])
    enc.registra_papel("C-1", "Clientes e ingresos", ruta_c1,
                       res_men.conclusion + f" Muestra MUS de {m.tamano} partidas "
                       f"(semilla {m.semilla}).", ["R001"], "concluido")
    enc.datos["fases"]["campo"] = "completa"

    # --- Fase 3: cierre e informe ---
    enc.añade_incorreccion(area="A", descripcion="Exceso de amortizacion en el elemento "
                                                 "ELE-9002 por no prorratear el alta",
                           efecto_resultado=-4800.0, epigrafe="8", tipo="factual",
                           corregida=False,
                           cualitativa="No afecta a covenants ni a la clasificacion de "
                                       "partidas; su efecto es inferior al umbral de "
                                       "insignificancia cualitativa.")
    enc.datos["informe"] = {
        "tipo_opinion": "favorable",
        "seccion_impuesto_sociedades": True,
        "variante_impuesto_sociedades": "no_obligada",
        "incertidumbre_empresa_en_funcionamiento": False,
        "revision_calidad_encargo": False,
    }
    enc.datos["fases"]["cierre"] = "completa"
    enc.guardar()

    rev = calidad.revisa(enc.datos)
    print(f"  Revision de calidad: {rev.datos['bloqueantes']} bloqueantes, "
          f"{rev.datos['a_resolver']} a resolver")

    fases_completas = all(v == "completa" for v in enc.datos["fases"].values())
    papeles = len(enc.datos["papeles"])
    comprueba("1", fases_completas and papeles >= 6,
              f"Las 4 fases se cierran dentro del plugin, con {papeles} papeles generados")
    riesgos_cubiertos = all(r["respuestas"] for r in enc.datos["riesgos"])
    comprueba("1", riesgos_cubiertos,
              f"Los {len(enc.datos['riesgos'])} riesgos tienen respuesta ejecutada")

    return {"encargo": enc, "sys": sys_df, "diario": diario_df, "meta": meta,
            "materialidad": mat, "epigrafes": epigrafes, "carpeta": carpeta,
            "leasing": (res_lea, resumen_lea, cuadros_lea)}


def _papel_area(ref: str, titulo_p: str, res, carpeta: str,
                detalles: dict, riesgos: list[str]) -> str:
    p = PapelDeTrabajo(ref, titulo_p, "TRANSPORTES SINTETICOS SL", EJERCICIO)
    p.alcance(res.concepto).riesgos(riesgos)
    p.concluye(res.conclusion, "CON EXCEPCIONES" if res.excepciones else "LIMPIO")
    for nombre, (df, tot) in detalles.items():
        p.detalle(df, nombre, tot)
    p.excepciones(res.excepciones)
    for k, v in res.datos.items():
        p.indicador(str(k).replace("_", " ").capitalize(), str(v)[:200])
    return p.guardar(os.path.join(carpeta, "01-papeles", f"{ref} {titulo_p[:30]}.xlsx"))


# ---------------------------------------------------------------------------
def criterio_2_leasings(ctx: dict) -> None:
    """2. 100 contratos heterogeneos -> cuadro completo con excepciones aisladas."""
    titulo("CRITERIO 2 - 100 contratos de arrendamiento heterogeneos")

    res, resumen, cuadros = ctx["leasing"]
    print(f"  {res.conclusion}")
    n = res.datos["contratos"]
    comprueba("2", n == 103,
              f"Se han ingerido {n} contratos de 3 ficheros con nomenclaturas distintas")
    comprueba("2", res.datos["procesados"] >= 100,
              f"{res.datos['procesados']} contratos procesados con cuadro completo")
    comprueba("2", not cuadros.empty and len(cuadros) > 3000,
              f"Cuadro de cuotas consolidado con {len(cuadros)} periodos")

    por_ej = leasing.cuadro_por_ejercicio(cuadros)
    comprueba("2", not por_ej.empty and {"cuotas", "carga_financiera",
                                         "amortizacion_capital"} <= set(por_ej.columns),
              f"Cuadro por ejercicio con cuota, carga financiera y capital "
              f"({len(por_ej)} ejercicios)")

    financieros = resumen[resumen["clasificacion"] == "FINANCIERO"]
    comprueba("2", len(financieros) > 0 and financieros["motivacion"].str.len().min() > 100,
              f"{len(financieros)} contratos clasificados como financieros con motivacion "
              "indicador a indicador")
    comprueba("2", (financieros["corriente_524"] + financieros["no_corriente_174"]
                    - financieros["deuda_viva_cierre"]).abs().max() < 0.02,
              "El reparto corriente/no corriente suma exactamente la deuda viva")

    # excepciones AISLADAS: los contratos con problema deliberado
    codigos = {e.codigo for e in res.excepciones}
    ids_exc = {e.descripcion.split()[1].rstrip(":") for e in res.excepciones
               if e.descripcion.startswith("Contrato ")}
    comprueba("2", "ARR-001" in codigos and "CS900" in ids_exc,
              "Aisla el contrato sin datos esenciales (CS900)")
    comprueba("2", "ARR-003" in codigos and "CS901" in ids_exc,
              "Aisla el contrato financieramente incoherente (CS901)")
    comprueba("2", "ARR-002" in codigos,
              f"Aisla los contratos con extraccion de baja confianza "
              f"({sum(1 for e in res.excepciones if e.codigo == 'ARR-002')})")
    proporcion = len(ids_exc) / n
    comprueba("2", proporcion < 0.20,
              f"Reporte por excepcion: solo {len(ids_exc)} de {n} contratos "
              f"({proporcion:.0%}) requieren atencion")

    # verificacion aritmetica independiente del tipo implicito
    fila = financieros.iloc[0]
    sub = cuadros[cuadros["contrato"] == fila["id"]]
    suma_capital = round(float(sub["amortizacion_capital"].sum()), 2)
    comprueba("2", abs(suma_capital - fila["importe_financiado"]) < 0.05,
              f"Recalculo aritmetico correcto: el capital amortizado del contrato "
              f"{fila['id']} suma su importe financiado ({suma_capital:,.2f} EUR)")


# ---------------------------------------------------------------------------
def criterio_3_comparador(ctx: dict) -> None:
    """3. El comparador detecta la diferencia introducida a proposito."""
    titulo("CRITERIO 3 - Diferencia deliberada entre memoria y balance")

    desgloses = json.load(open(os.path.join(FIXT, "memoria_desgloses.json"),
                               encoding="utf-8"))
    estados = json.load(open(os.path.join(FIXT, "estados_financieros.json"),
                             encoding="utf-8"))
    res = comparador.memoria_vs_estados(desgloses, estados)
    print(f"  {res.conclusion}")
    for e in res.ordenadas():
        print(f"    -> {e.descripcion}")

    detectada = any(abs((e.importe or 0) - 4850.00) < 0.01 for e in res.excepciones)
    comprueba("3", detectada,
              "Detecta la diferencia de 4.850,00 EUR entre la memoria y el balance")
    comprueba("3", len(res.excepciones) == 1,
              f"Reporta solo la diferencia, no los {len(estados)} conceptos comparados")

    # checklist de desgloses: nota ausente y nota heredada
    texto = open(os.path.join(FIXT, "memoria_2025.txt"), encoding="utf-8").read()
    anterior = open(os.path.join(FIXT, "memoria_2024.txt"), encoding="utf-8").read()
    res_chk = comparador.checklist_memoria(texto, "PYME",
                                           {"arrendamientos": True}, anterior)
    print(f"  {res_chk.conclusion}")
    ausentes = [e for e in res_chk.excepciones if e.codigo == "MEM-001"]
    heredadas = [e for e in res_chk.excepciones if e.codigo == "MEM-002"]
    comprueba("3", len(ausentes) > 0,
              f"Detecta {len(ausentes)} desgloses obligatorios ausentes de la memoria")
    comprueba("3", len(heredadas) > 0,
              f"Detecta {len(heredadas)} parrafos con cifras heredados sin actualizar")

    # informe vs cuentas definitivas
    res_inf = comparador.informe_vs_ccaa_definitivas(
        {"denominacion_social": "TRANSPORTES SINTETICOS SL", "nif": "B00000000",
         "ejercicio": 2025, "fecha_cierre": "2025-12-31", "marco": "PGC PYMES",
         "modelo": "PYME", "fecha_formulacion": "2026-03-31", "total_activo": 900000.0},
        {"denominacion_social": "TRANSPORTES SINTETICOS SL", "nif": "B00000000",
         "ejercicio": 2024, "fecha_cierre": "2025-12-31", "marco": "PGC PYMES",
         "modelo": "PYME", "fecha_formulacion": "2026-03-31", "total_activo": 912500.0})
    bloqueantes = [e for e in res_inf.excepciones if e.severidad == "BLOQUEANTE"]
    comprueba("3", len(bloqueantes) == 2,
              f"La verificacion previa a la firma detecta {len(bloqueantes)} discrepancias "
              "bloqueantes entre informe y cuentas definitivas (ejercicio y total activo)")


# ---------------------------------------------------------------------------
def criterio_4_trazabilidad(ctx: dict) -> None:
    """4. Todos los calculos son reproducibles por script y trazables al origen."""
    titulo("CRITERIO 4 - Reproducibilidad y trazabilidad")

    # (a) reproducibilidad: misma entrada, misma salida
    sys_a, _ = ingesta.normaliza_sumas_y_saldos(
        os.path.join(FIXT, "sumas_y_saldos_2025.xlsx"))
    sys_b, _ = ingesta.normaliza_sumas_y_saldos(
        os.path.join(FIXT, "sumas_y_saldos_2025.xlsx"))
    comprueba("4", sys_a.equals(sys_b), "La ingesta es determinista (dos ejecuciones "
                                        "producen exactamente el mismo DataFrame)")

    m1 = muestreo.mus(ctx["diario"][ctx["diario"]["cuenta"].str.startswith("43")],
                      "debe", ctx["materialidad"].ejecucion, semilla=12345)
    m2 = muestreo.mus(ctx["diario"][ctx["diario"]["cuenta"].str.startswith("43")],
                      "debe", ctx["materialidad"].ejecucion, semilla=12345)
    comprueba("4", m1.seleccion.index.tolist() == m2.seleccion.index.tolist()
              and m1.semilla == m2.semilla == 12345,
              f"El muestreo es reproducible con la semilla registrada ({m1.semilla})")

    mat1 = materialidad.calcula({"cifra_negocios": 1_850_000.0}, "ESTANDAR")
    mat2 = materialidad.calcula({"cifra_negocios": 1_850_000.0}, "ESTANDAR")
    comprueba("4", mat1.global_ == mat2.global_ and mat1.ejecucion == mat2.ejecucion,
              f"La materialidad es determinista (MG {mat1.global_:,.2f} EUR)")

    # (b) trazabilidad a celda concreta
    meta = ctx["meta"]
    t = ingesta.traza_de(meta, 12, "saldo")
    comprueba("4", "!" in str(t) and any(c.isdigit() for c in str(t)),
              f"Traza a celda concreta del fichero de origen: {t}")

    reg = RegistroTrazas()
    reg.anota("Total debe", 100.0, Traza(meta["fichero"], meta["hoja"], "C7"))
    reg.pendiente("Valor razonable del bien", "contratos.xlsx", "no consta en el contrato")
    reg.juicio("Vida util economica", "contratos.xlsx", "requiere criterio del auditor")
    filas = reg.filas()
    comprueba("4", len(filas) == 3 and len(reg.pendientes) == 2,
              "Los datos ausentes se marcan [PENDIENTE-CLIENTE] / [JUICIO-AUDITOR], "
              "nunca se estiman en silencio")

    # (c) huella de los ficheros fuente
    h = huella(os.path.join(FIXT, "sumas_y_saldos_2025.xlsx"))
    comprueba("4", len(h) == 64,
              f"Los ficheros fuente quedan registrados con su SHA-256 ({h[:16]}...)")

    # (d) el papel de trabajo lleva la hoja de traza
    from openpyxl import load_workbook
    papel = [p for p in ctx["encargo"].datos["papeles"] if p["ref"] == "2.1"][0]
    wb = load_workbook(papel["fichero"])
    comprueba("4", {"Conclusion", "Traza", "Excepciones"} <= set(wb.sheetnames),
              f"El papel de trabajo tiene las hojas {wb.sheetnames}")

    # (e) formulas vivas en las lineas de total
    ws = wb["Sumas y saldos"]
    formulas = [ws.cell(row=ws.max_row, column=c).value
                for c in range(1, ws.max_column + 1)]
    comprueba("4", any(isinstance(v, str) and v.startswith("=SUM(") for v in formulas),
              "La fila de totales lleva formula viva (=SUM), revisable sin reejecutar")


# ---------------------------------------------------------------------------
def criterio_5_revision_calidad() -> None:
    """5. La revision detecta un papel sin conclusion y un riesgo sin respuesta."""
    titulo("CRITERIO 5 - Revision de calidad sobre un archivo con defectos")

    enc = Encargo.abrir(os.path.join(FIXT, "encargo_demo"))
    res = calidad.revisa(enc.datos)
    print()
    print(calidad.panel_del_socio(enc.datos, res))
    print()

    sin_respuesta = [e for e in res.excepciones if e.codigo == "CAL-001"]
    sin_conclusion = [e for e in res.excepciones if e.codigo == "CAL-010"]

    comprueba("5", any("R003" in e.descripcion for e in sin_respuesta),
              "Detecta el riesgo R003 sin ningun procedimiento que lo responda")
    comprueba("5", any(e.severidad == "BLOQUEANTE" for e in sin_respuesta),
              "Lo clasifica como BLOQUEANTE por tratarse de un riesgo significativo")
    comprueba("5", any("F-1" in e.descripcion for e in sin_conclusion),
              "Detecta el papel F-1 concluido sin conclusion redactada")
    comprueba("5", not res.datos["puede_firmarse"],
              "Concluye que el archivo NO esta en condiciones de firma")

    # el panel del socio cabe en una pagina
    panel = calidad.panel_del_socio(enc.datos, res)
    lineas = len(panel.splitlines())
    comprueba("5", lineas <= 60,
              f"El panel del socio ocupa {lineas} lineas (cabe en una pagina)")

    # modo pre-vuelo: mismo archivo, menos alarmas prematuras
    res_pv = calidad.revisa(enc.datos, pre_vuelo=True)
    comprueba("5", res_pv.datos["bloqueantes"] <= res.datos["bloqueantes"],
              f"El modo pre-vuelo degrada las alarmas prematuras "
              f"({res.datos['bloqueantes']} -> {res_pv.datos['bloqueantes']} bloqueantes)")

    # archivo correcto: no debe generar falsos positivos criticos
    limpio = json.loads(json.dumps(enc.datos))
    limpio["riesgos"][2]["respuestas"] = ["E-1"]
    limpio["papeles"][1]["conclusion"] = ("Se ha recalculado la totalidad de los "
                                          "contratos de arrendamiento y conciliado con la "
                                          "contabilidad sin diferencias significativas.")
    limpio["papeles"][1]["estado"] = "concluido"
    limpio["papeles"].append({"ref": "E-1", "titulo": "Financiacion",
                              "fichero": "01-papeles/E-1.xlsx",
                              "conclusion": "La deuda financiera recalculada cuadra con "
                                            "la contabilidad y las garantias estan "
                                            "desglosadas en memoria.",
                              "riesgos": ["R003"], "estado": "concluido"})
    limpio["fases"] = {k: "completa" for k in limpio["fases"]}
    limpio["informe"] = {"tipo_opinion": "favorable",
                         "seccion_impuesto_sociedades": True}
    res_limpio = calidad.revisa(limpio)
    comprueba("5", res_limpio.datos["bloqueantes"] == 0,
              f"Sobre el archivo corregido no quedan bloqueantes "
              f"({res_limpio.datos['a_resolver']} pendientes de documentacion)")


# ---------------------------------------------------------------------------
def criterio_6_informe() -> None:
    """6. El modelo de informe corresponde a la version normativa vigente."""
    titulo("CRITERIO 6 - Modelo de informe conforme a la RICAC de 22/01/2026")

    ruta = os.path.join(RAIZ, "shared", "templates", "informe-auditoria.md")
    texto = open(ruta, encoding="utf-8").read()

    comprueba("6", "22 de enero de 2026" in texto and "BOE-A-2026-2234" in texto,
              "El modelo se identifica con la Resolucion del ICAC de 22/01/2026")
    comprueba("6", "Informe relativo al impuesto sobre sociedades" in texto,
              "Incluye la seccion relativa al impuesto sobre sociedades")
    bajo = texto.lower()
    comprueba("6", "informe sobre otros requerimientos legales y reglamentarios" in bajo
              and bajo.index("informe sobre otros requerimientos legales") <
              bajo.index("informe relativo al impuesto sobre sociedades"),
              "La ubica dentro de 'Informe sobre otros requerimientos legales y "
              "reglamentarios'")
    comprueba("6", "disposicion adicional undecima" in texto.lower()
              or "disposición adicional undécima" in texto.lower(),
              "La fundamenta en la disposicion adicional undecima de la LAC")
    comprueba("6", "750 millones" in texto,
              "Recoge el umbral de 750 millones de euros de cifra de negocios consolidada")
    comprueba("6", "22/06/2025" in texto and "01/01/2026" in texto,
              "Distingue las dos puertas de aplicacion (22/06/2025 y 01/01/2026)")
    for variante in ("no estaba obligada", "ha sido publicado", "no ha sido publicado"):
        comprueba("6", variante in texto,
                  f"Incluye la variante de redaccion: '{variante}'")
    for opinion in ("Opinión con salvedades", "Opinión desfavorable", "Opinión denegada"):
        comprueba("6", opinion in texto, f"Incluye el modelo de {opinion.lower()}")
    comprueba("6", "Incertidumbre material relacionada con la empresa en funcionamiento"
              in texto, "Incluye la seccion de incertidumbre sobre empresa en funcionamiento")
    comprueba("6", "[VERIFICAR-LITERAL-ICAC]" in texto,
              "Marca expresamente los parrafos pendientes de contraste literal con el "
              "PDF oficial del ICAC")


# ---------------------------------------------------------------------------
def main() -> int:
    os.makedirs(SALIDA, exist_ok=True)
    if not os.path.exists(os.path.join(FIXT, "sumas_y_saldos_2025.xlsx")):
        print("Generando fixtures...")
        import generar_fixtures
        generar_fixtures.main()

    ctx = criterio_1_encargo_completo()
    criterio_2_leasings(ctx)
    criterio_3_comparador(ctx)
    criterio_4_trazabilidad(ctx)
    criterio_5_revision_calidad()
    criterio_6_informe()

    titulo("RESUMEN")
    por_criterio: dict[str, list[bool]] = {}
    for crit, ok, _ in RESULTADOS:
        por_criterio.setdefault(crit, []).append(ok)
    nombres = {
        "1": "Un encargo pequeno se recorre entero sin salir del plugin",
        "2": "100 contratos heterogeneos -> cuadro completo con excepciones aisladas",
        "3": "El comparador detecta la diferencia deliberada memoria/balance",
        "4": "Calculos reproducibles por script y trazables al fichero de origen",
        "5": "La revision detecta papel sin conclusion y riesgo sin respuesta",
        "6": "El modelo de informe corresponde a la version normativa vigente",
    }
    todo_ok = True
    for c in sorted(por_criterio):
        oks = por_criterio[c]
        ok = all(oks)
        todo_ok &= ok
        print(f"  [{'OK  ' if ok else 'FALLA'}] Criterio {c}: {sum(oks)}/{len(oks)} "
              f"comprobaciones - {nombres[c]}")
    fallidas = [d for _, ok, d in RESULTADOS if not ok]
    if fallidas:
        print("\n  Comprobaciones fallidas:")
        for f in fallidas:
            print(f"    - {f}")
    print()
    print(f"TOTAL: {sum(1 for _, ok, _ in RESULTADOS if ok)}/{len(RESULTADOS)} "
          f"comprobaciones superadas.")
    return 0 if todo_ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
